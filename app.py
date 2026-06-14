# app.py  --  Streamlit UI for the Lab Results Analyzer
# Run: py -3 -m streamlit run app.py
import sys, os, io, re, collections, socket
from datetime import date

# ── add soil_lab_tool to path ─────────────────────────────────────
ROOT     = os.path.dirname(os.path.abspath(__file__))
TOOL_DIR = os.path.join(ROOT, 'soil_lab_tool')
if TOOL_DIR not in sys.path:
    sys.path.insert(0, TOOL_DIR)

THRESH_DIR = os.path.join(TOOL_DIR, 'thresholds')
LAB_DIR    = os.path.join(ROOT, 'Laboratory_results')

# ── company logo (base64 embed) ────────────────────────────────────
import base64, pathlib
_LOGO_PATH = pathlib.Path(__file__).parent / "assets" / "adama_logo.png"
if _LOGO_PATH.exists():
    _LOGO_B64 = base64.b64encode(_LOGO_PATH.read_bytes()).decode()
    LOGO_TAG = f'<img src="data:image/png;base64,{_LOGO_B64}" style="height:48px;">'
else:
    # fallback to legacy logo names
    def _logo_b64_fallback() -> str:
        for name in ('תמונה1.png', 'logo.png'):
            p = os.path.join(ROOT, name)
            if os.path.exists(p):
                with open(p, 'rb') as _f:
                    return base64.b64encode(_f.read()).decode()
        return ""
    _LOGO_B64 = _logo_b64_fallback()
    if _LOGO_B64:
        LOGO_TAG = f'<img src="data:image/png;base64,{_LOGO_B64}" style="height:48px;">'
    else:
        LOGO_TAG = '<span style="font-size:1.4rem;font-weight:800;color:#1e3a4f;">אדמה</span>'

LOGO_B64 = _LOGO_B64   # kept for any legacy references

import streamlit as st
import pandas as pd

# ── page config ───────────────────────────────────────────────────
st.set_page_config(
    page_title="מערכת ניתוח תוצאות מעבדה",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── detect LAN IP for sharing ─────────────────────────────────────
def _local_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "localhost"

LAN_IP  = _local_ip()
APP_URL = f"http://{LAN_IP}:8501"


def _pid_norm(name: str) -> str:
    """Normalize a borehole name for PID matching: remove dashes and spaces."""
    return re.sub(r'[-\s]', '', name).strip()


def _parse_pid_file(uploaded_file) -> dict:
    """Parse a PID field-data file."""
    df = pd.read_excel(uploaded_file, header=None)
    df.iloc[:, 0] = df.iloc[:, 0].ffill()
    pid_data: dict = {}
    for bh_raw, grp in df.groupby(df.columns[0], sort=False):
        bh = str(bh_raw).strip()
        if not bh or bh.lower() in ("nan", "none"):
            continue
        key = _pid_norm(bh)
        if not key:
            continue
        entries: list = []
        for _, row in grp.iterrows():
            try:
                depth_to = float(row.iloc[2])
            except (ValueError, TypeError, IndexError):
                continue
            pid_raw = row.iloc[7] if row.shape[0] > 7 else None
            if pid_raw is None:
                continue
            if isinstance(pid_raw, float) and pd.isna(pid_raw):
                continue
            pid_str = str(pid_raw).strip().upper()
            if pid_str in ("N/A", "NA", ""):
                continue
            try:
                pid_val = float(pid_raw)
            except (ValueError, TypeError):
                continue
            entries.append((depth_to, pid_val))
        if entries:
            pid_data[key] = entries
    return pid_data


# ══════════════════════════════════════════════════════════════════
# CSS — Adama design system
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Heebo:wght@300;400;500;600;700;800;900&display=swap');

* { box-sizing: border-box; font-family: 'Heebo', sans-serif !important; }
html, body {
    direction: ltr;
    font-family: 'Heebo', sans-serif !important;
    font-size: 18px;
    background: #f0f4f8;
    margin: 0;
}
[data-testid="stSidebarContent"], [data-testid="stMain"], .block-container { direction: rtl; }

p, label, div, span, input, select { font-size: inherit !important; }

/* Hide Streamlit chrome */
#MainMenu, footer, header { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }
.stDeployButton { display: none !important; }

/* Main container */
[data-testid="stMain"] {
    background: #f0f4f8;
    padding: 0 !important;
}
[data-testid="stMain"] .block-container {
    direction: rtl;
    padding: 0 !important;
    max-width: 100% !important;
}

/* Top navigation bar */
.nav-bar {
    background: #1e3a4f;
    padding: 0 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    height: 72px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.15);
    position: sticky;
    top: 0;
    z-index: 999;
}
.nav-logo { display: flex; align-items: center; gap: 20px; }
.nav-links { display: flex; gap: 4px; }
.nav-link {
    color: #94b8c8 !important;
    padding: 10px 20px;
    border-radius: 6px;
    font-size: 1.1rem;
    font-weight: 500;
    cursor: pointer;
    transition: all 0.2s;
    text-decoration: none !important;
    border: none;
    background: transparent;
}
.nav-link:hover { background: rgba(255,255,255,0.1); color: white !important; }
.nav-link.active {
    background: #4a7a8a;
    color: white !important;
    font-weight: 600;
}

/* Page wrapper */
.page-wrapper {
    max-width: 1280px;
    margin: 0 auto;
    padding: 2rem;
}

/* Cards */
.card {
    background: white;
    border-radius: 12px;
    padding: 1.5rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08), 0 4px 12px rgba(0,0,0,0.05);
    margin-bottom: 1.25rem;
    border: 1px solid #e8eef3;
}
.card-header {
    font-size: 1.2rem;
    font-weight: 700;
    color: #1e3a4f;
    margin-bottom: 1rem;
    padding-bottom: 0.75rem;
    border-bottom: 2px solid #f0f4f8;
}

/* Hero section */
.hero {
    background: linear-gradient(135deg, #1e3a4f 0%, #2d5a6e 50%, #4a7a8a 100%);
    border-radius: 16px;
    padding: 3rem 2.5rem;
    color: white;
    margin-bottom: 2rem;
}
.hero h1 { font-size: 3.2rem; font-weight: 900; margin: 0 0 0.5rem; }
.hero p { font-size: 1.3rem; opacity: 0.85; margin: 0; font-weight: 300; }

/* Stat boxes */
.stat-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem; margin-bottom: 1.5rem; }
.stat-box {
    background: white;
    border-radius: 12px;
    padding: 1.25rem;
    text-align: center;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    border-top: 3px solid #4a7a8a;
}
.stat-number { font-size: 2.8rem; font-weight: 800; color: #1e3a4f; }
.stat-label { font-size: 1rem; color: #64748b; font-weight: 500; margin-top: 4px; }

/* Step indicator */
.steps { display: flex; gap: 1rem; margin-bottom: 1.5rem; }
.step {
    flex: 1;
    background: white;
    border-radius: 12px;
    padding: 1.25rem;
    text-align: center;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    border: 1px solid #e8eef3;
}
.step-num {
    width: 36px; height: 36px;
    background: #1e3a4f;
    color: white;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-weight: 700; font-size: 1rem;
    margin: 0 auto 0.75rem;
}
.step-title { font-weight: 700; color: #1e3a4f; font-size: 1.1rem; }
.step-desc { font-size: 0.95rem; color: #64748b; margin-top: 4px; }

/* Badge tags */
.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    margin: 2px;
}
.badge-blue { background: #dbeafe; color: #1e40af; }
.badge-green { background: #dcfce7; color: #166534; }
.badge-gray { background: #f1f5f9; color: #475569; }

/* Upload area */
.upload-card {
    border: 2px dashed #cbd5e1;
    border-radius: 12px;
    padding: 2rem;
    text-align: center;
    background: #f8fafc;
    transition: all 0.2s;
}
.upload-card:hover { border-color: #4a7a8a; background: #f0f7fa; }

/* Download buttons */
.dl-btn {
    background: #1e3a4f;
    color: white;
    border: none;
    border-radius: 8px;
    padding: 0.75rem 1.5rem;
    font-size: 1rem;
    font-weight: 600;
    cursor: pointer;
    width: 100%;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    transition: background 0.2s;
}
.dl-btn:hover { background: #2d5a6e; }

/* Progress bar */
.progress-wrap {
    background: #e2e8f0;
    border-radius: 999px;
    height: 6px;
    overflow: hidden;
    margin: 1rem 0;
}
.progress-fill {
    height: 100%;
    background: linear-gradient(90deg, #4a7a8a, #7a9a7a);
    border-radius: 999px;
    transition: width 0.4s ease;
}

/* Table styling */
.lab-table { width: 100%; border-collapse: collapse; font-size: 1rem; }
.lab-table th {
    background: #1e3a4f;
    color: white;
    padding: 10px 14px;
    text-align: right;
    font-weight: 600;
}
.lab-table td {
    padding: 10px 14px;
    border-bottom: 1px solid #f1f5f9;
    color: #374151;
}
.lab-table tr:hover td { background: #f8fafc; }

/* Footer */
.footer {
    background: #1e3a4f;
    color: #94b8c8;
    text-align: center;
    padding: 1.5rem;
    font-size: 0.95rem;
    margin-top: 3rem;
}
.footer strong { color: white; }


section[data-testid="stMain"] {
    margin-right: 0 !important;
    padding-right: 0 !important;
    width: 100% !important;
}

/* RTL typography for Streamlit widgets */
.stMarkdown p, .stMarkdown li,
.stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4 {
    direction: rtl; text-align: right;
}
.stTextInput label, .stSelectbox label,
.stMultiSelect label, .stFileUploader label,
.stCheckbox label, .stRadio label,
.stMetric label, .stMetric div,
[data-testid="stMetricLabel"], [data-testid="stMetricValue"],
[data-testid="stCaptionContainer"] {
    direction: rtl; text-align: right;
}

/* section-card for soil page */
.section-card {
    background: #ffffff;
    border: 1px solid #e2eef3;
    border-radius: 12px;
    padding: 1.5rem;
    margin: 0 2rem 1rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
}

/* banners */
.info-banner {
    background: #eff6ff;
    border: 1px solid #bfdbfe;
    border-radius: 8px;
    padding: 0.75rem 1rem;
    direction: rtl;
    font-size: 0.875rem;
    color: #1e40af;
    margin-bottom: 0.75rem;
    border-right: 3px solid #3b82f6;
}
.success-banner {
    background: #f0fdf4;
    border: 1px solid #86efac;
    border-radius: 8px;
    padding: 0.75rem 1rem;
    direction: rtl;
    font-size: 0.875rem;
    color: #15803d;
    margin-bottom: 0.75rem;
    border-right: 3px solid #16a34a;
}

/* analysis-type pill badges */
.type-badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 999px;
    font-size: 0.8rem;
    font-weight: 600;
    letter-spacing: 0.2px;
    margin: 2px 3px;
    color: white;
}

/* sidebar label */
.sidebar-label {
    font-size: 0.65rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #64748b;
    margin: 1rem 0 0.3rem;
    padding-bottom: 0.2rem;
    border-bottom: 1px solid #334155;
}

/* upload zone */
[data-testid="stFileUploader"] {
    border: 2px dashed #cbd5e1;
    border-radius: 8px;
    padding: 0.5rem;
    background: #f8fafc;
    transition: border-color 0.2s;
}
[data-testid="stFileUploader"]:hover {
    border-color: #4a7a8a;
    background: #f0f7fa;
}
[data-testid="stFileUploader"] label {
    font-size: 0.95rem !important;
    font-weight: 600 !important;
    color: #334155 !important;
}

/* download button */
.stDownloadButton button {
    width: 100%;
    background: linear-gradient(135deg, #1e3a4f, #2d5a6e);
    color: white;
    font-weight: 700;
    font-size: 0.95rem;
    border-radius: 8px;
    border: none;
    padding: 0.65rem;
}
.stDownloadButton button:hover {
    background: linear-gradient(135deg, #2d5a6e, #4a7a8a);
}

/* metric cards */
[data-testid="metric-container"] {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    padding: 0.75rem 1rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
[data-testid="stMetricValue"] {
    font-size: 1.6rem !important;
    color: #1e3a4f !important;
    font-weight: 700 !important;
}
[data-testid="stMetricLabel"] {
    font-size: 0.8rem !important;
    color: #64748b !important;
}

/* step pills (soil page) */
.step-row { display: flex; gap: 0.75rem; margin-bottom: 1.25rem; direction: rtl; }
.step-pill {
    display: flex; align-items: center; gap: 0.4rem;
    background: #f1f5f9; border: 1.5px solid #cbd5e1; border-radius: 8px;
    padding: 0.3rem 0.9rem; font-size: 0.8rem; color: #64748b;
    font-weight: 500; flex: 1; justify-content: center;
}
.step-pill.active { background: #e8f4f8; border-color: #4a7a8a; color: #1e3a4f; font-weight: 700; }
.step-pill.done { background: #f0fdf4; border-color: #22c55e; color: #15803d; }

/* threshold pill toggles */
div:has(.thresh-pill-marker) + div [data-testid="stCheckbox"] label {
    display: inline-flex !important;
    align-items: center !important;
    gap: 6px !important;
    padding: 5px 18px !important;
    border-radius: 999px !important;
    border: 1.5px solid #cbd5e1 !important;
    background: #f1f5f9 !important;
    font-size: 0.85rem !important;
    font-weight: 600 !important;
    color: #64748b !important;
    cursor: pointer !important;
    transition: background .15s, border-color .15s, color .15s !important;
    user-select: none !important;
}
div:has(.thresh-pill-marker) + div [data-testid="stCheckbox"]:has(input:checked) label {
    background: #e8f4f8 !important;
    border-color: #4a7a8a !important;
    color: #1e3a4f !important;
}
div:has(.thresh-pill-marker) + div [data-testid="stCheckbox"] label svg {
    display: none !important;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# LOAD THRESHOLD MANAGER
# ══════════════════════════════════════════════════════════════════
@st.cache_resource(show_spinner="טוען ערכי סף...")
def load_threshold_manager(_mtime):
    from core.threshold_manager import ThresholdManager
    MAIN_THRESH  = os.path.join(THRESH_DIR, 'soil_vsl_tier1_v7_2024.xlsx')
    VSL_FULL     = os.path.join(THRESH_DIR, 'soil_vsl_v7_full.xlsx')
    PFAS_THRESH     = os.path.join(LAB_DIR,   'נספח לטבלת ערכי סף - PFAS.xlsx')
    PFAS_THRESH_ALT = os.path.join(THRESH_DIR, 'pfas_thresholds.xlsx')
    vsl_full_path = VSL_FULL if os.path.exists(VSL_FULL) else None
    pfas_path = (PFAS_THRESH     if os.path.exists(PFAS_THRESH)
                 else PFAS_THRESH_ALT if os.path.exists(PFAS_THRESH_ALT)
                 else None)
    return ThresholdManager(MAIN_THRESH, pfas_path=pfas_path, vsl_full_path=vsl_full_path)


# ══════════════════════════════════════════════════════════════════
# MODULE IMPORTS
# ══════════════════════════════════════════════════════════════════
try:
    from core.excel_output import LabReportExcel
    from core.word_output  import (LabReportWord, parse_als_file, build_word_report,
                                   load_threshold_file, get_tier1_col, tier1_label)
    from core.xrf_output   import build_xrf_simple_excel
    from parsers import get_parser, auto_detect_category, auto_detect_lab
    _tm_py    = os.path.join(TOOL_DIR, 'core', 'threshold_manager.py')
    _tm_mtime = os.path.getmtime(_tm_py)
    tm = load_threshold_manager(_tm_mtime)
except Exception as e:
    st.error(f"שגיאת טעינת מודולים: {e}")
    st.stop()


# ══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════
def _build_pivot_table(recs):
    """Build a compound × sample_id pivot DataFrame for preview."""
    rows = []
    for r in recs:
        val  = r.get('value')
        flag = r.get('flag', '')
        loq  = r.get('loq')
        if flag in ('ND', '<LOD', '<LOQ') and loq is not None:
            display_val = f"<{loq}"
        elif val is not None:
            display_val = f"{val:.4g}" if isinstance(val, (int, float)) else str(val)
        else:
            display_val = flag or ''
        rows.append({
            'תרכובת':    r.get('compound', ''),
            'sample_id': r.get('sample_id', ''),
            'value':     display_val,
        })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    try:
        pivot = df.pivot_table(index='תרכובת', columns='sample_id',
                               values='value', aggfunc='first')
        pivot.columns.name = None
        return pivot
    except Exception:
        return df[['תרכובת', 'sample_id', 'value']]


_PREVIEW_LEGEND = (
    '<div style="font-size:0.75rem;direction:rtl;margin-top:6px;display:flex;'
    'gap:10px;align-items:center;">'
    '<span style="background:#FFFF00;padding:2px 10px;border-radius:2px;'
    'border:1px solid #ccc;">מעל VSL</span>'
    '<span style="background:#FFC000;padding:2px 10px;border-radius:2px;'
    'border:1px solid #ccc;">מעל Tier 1</span>'
    '</div>'
)


def _build_styled_pivot(recs: list[dict], tm, selected_thresholds: list[str]):
    """Return (Styler, has_any_color) for a compound × sample_id pivot."""
    if not recs:
        return pd.DataFrame().style, False

    rows_d, rows_n, cas_map = [], [], {}
    for r in recs:
        val      = r.get('value')
        flag     = r.get('flag', '')
        loq      = r.get('loq')
        compound = r.get('compound', '')
        sample   = r.get('sample_id', '')
        cas      = r.get('cas', '')

        if flag in ('ND', '<LOD', '<LOQ') and loq is not None:
            disp = f"<{loq}"
        elif val is not None:
            disp = f"{val:.4g}" if isinstance(val, (int, float)) else str(val)
        else:
            disp = flag or ''

        rows_d.append({'תרכובת': compound, '_s': sample, '_v': disp})
        rows_n.append({'תרכובת': compound, '_s': sample,
                       '_v': val if isinstance(val, (int, float)) else None})
        cas_map.setdefault(compound, cas)

    try:
        piv_d = (pd.DataFrame(rows_d)
                 .pivot_table(index='תרכובת', columns='_s', values='_v', aggfunc='first'))
        piv_n = (pd.DataFrame(rows_n)
                 .pivot_table(index='תרכובת', columns='_s', values='_v', aggfunc='first'))
        piv_d.columns.name = None
        piv_n.columns.name = None
    except Exception:
        return pd.DataFrame(rows_d).style, False

    vsl_keys   = [k for k in selected_thresholds if 'VSL' in k]
    tier1_keys = [k for k in selected_thresholds if k not in vsl_keys]
    has_colors = bool((vsl_keys or tier1_keys) and tm is not None)

    colors = pd.DataFrame('', index=piv_d.index, columns=piv_d.columns)

    if has_colors:
        for compound in piv_d.index:
            cas = cas_map.get(compound, '')
            for sample in piv_d.columns:
                try:
                    raw = piv_n.loc[compound, sample]
                    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                        continue
                    num_val = float(raw)
                except (TypeError, ValueError, KeyError):
                    continue

                for tk in tier1_keys:
                    thresh = tm.get_threshold_with_name(cas, tk, compound)
                    if thresh is not None and num_val > thresh:
                        colors.loc[compound, sample] = 'background-color: #FFC000'
                        break

                if colors.loc[compound, sample]:
                    continue

                for vk in vsl_keys:
                    thresh = tm.get_threshold_with_name(cas, vk, compound)
                    if thresh is not None and num_val > thresh:
                        colors.loc[compound, sample] = 'background-color: #FFFF00'
                        break

    styled = piv_d.style.apply(lambda _: colors, axis=None)
    return styled, has_colors


def _steps(step: int):
    labels = ["① העלאת קובץ", "② בחירת ערכי סף", "③ הורדת דוח"]
    pills = ""
    for i, lbl in enumerate(labels, 1):
        cls = "active" if i == step else ("done" if i < step else "step-pill")
        icon = "✅ " if i < step else ""
        pills += f'<div class="step-pill {cls}">{icon}{lbl}</div>'
    st.markdown(f'<div class="step-row">{pills}</div>', unsafe_allow_html=True)


_PREV_LABELS = {
    "SOIL_VOC": "VOC", "SOIL_SVOC": "SVOC", "SOIL_TPH": "TPH",
    "SOIL_PFAS": "PFAS", "SOIL_METALS": "מתכות",
    "GW_VOC": "מי תהום VOC", "SOIL_GAS_VOC": "גז קרקע", "GAS_VOC": "גז קרקע",
}


@st.dialog("📊 תצוגה מקדימה", width="large")
def _preview_dialog(records, tm, project_name, client_name, rep_date,
                    selected_thresholds, combine_tph_voc, combine_tph_mbtex, pid_map):
    import openpyxl as _oxl

    buf = io.BytesIO()
    with st.spinner("בונה תצוגה מקדימה..."):
        try:
            LabReportExcel(
                records             = records,
                threshold_manager   = tm,
                output_path         = buf,
                project_name        = project_name,
                client              = client_name,
                report_date         = rep_date,
                selected_thresholds = selected_thresholds,
                combine_tph_voc     = combine_tph_voc,
                combine_tph_mbtex   = combine_tph_mbtex,
                pid_map             = pid_map,
            ).build()
        except Exception as _e:
            st.error(f"שגיאה בבניית תצוגה מקדימה: {_e}")
            return

    buf.seek(0)
    wb = _oxl.load_workbook(buf, data_only=True)
    if not wb.sheetnames:
        st.info("אין נתונים להצגה")
        return

    tabs = st.tabs(wb.sheetnames)
    for tab, sname in zip(tabs, wb.sheetnames):
        with tab:
            ws = wb[sname]

            for mr in list(ws.merged_cells.ranges):
                top_val  = ws.cell(mr.min_row, mr.min_col).value
                ws.unmerge_cells(str(mr))
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        ws.cell(r, c).value = top_val

            all_rows = list(ws.iter_rows())
            if not all_rows:
                st.info("גיליון ריק")
                continue

            n_cols = max(len(row) for row in all_rows)
            data_list, css_list = [], []
            for row in all_rows:
                row_vals, row_css = [], []
                for cell in row:
                    v = cell.value
                    row_vals.append("" if v is None else
                                    (v if isinstance(v, (int, float)) else str(v)))
                    css_bg = ""
                    try:
                        f = cell.fill
                        if getattr(f, 'fill_type', None) == "solid":
                            fc = getattr(f, 'fgColor', None)
                            if fc and getattr(fc, 'type', None) == 'rgb':
                                rgb = fc.rgb
                                hx = rgb[2:] if len(rgb) == 8 else rgb
                                if hx.upper() not in ("FFFFFF", "000000", "00000000"):
                                    css_bg = f"background-color: #{hx}; color: #000;"
                    except Exception:
                        pass
                    row_css.append(css_bg)
                pad = n_cols - len(row_vals)
                row_vals += [""] * pad
                row_css  += [""] * pad
                data_list.append(row_vals)
                css_list.append(row_css)

            df = pd.DataFrame(data_list, columns=list(range(n_cols)))

            n_dr, n_dc = len(df), len(df.columns)
            padded = []
            for ri in range(n_dr):
                row = css_list[ri] if ri < len(css_list) else []
                padded.append([row[ci] if ci < len(row) else "" for ci in range(n_dc)])
            style_df = pd.DataFrame(padded, index=df.index, columns=df.columns)

            styled = (df.style
                      .apply(lambda _: style_df, axis=None)
                      .hide(axis='columns')
                      .hide(axis='index'))
            st.dataframe(styled, use_container_width=True, height=480)


# ══════════════════════════════════════════════════════════════════
# ROUTING
# ══════════════════════════════════════════════════════════════════
page = st.query_params.get("page", "home")


# ── Navigation bar (HTML anchor tags, same-tab routing) ──────────
def _render_nav(active: str):
    _NAV_PAGES = [
        ("soil",        "📊 ניתוח נתונים"),
        ("groundwater", "💧 דוחות מי תהום"),
        ("guide",       "📖 מדריך"),
    ]
    logo_html = (
        f'<img src="data:image/png;base64,{LOGO_B64}" '
        f'style="height:44px;vertical-align:middle;background:white;border-radius:6px;padding:4px 8px;">'
        if LOGO_B64 else '<strong style="color:white;font-size:1.4rem;">אדמה</strong>'
    )
    links_html = ""
    for p, lbl in _NAV_PAGES:
        cls = "nav-link active" if p == active else "nav-link"
        links_html += f'<a href="?page={p}" class="{cls}">{lbl}</a>'
    st.markdown(
        f'<div class="nav-bar">'
        f'<div class="nav-logo">'
        f'{logo_html}'
        f'<span style="color:rgba(255,255,255,0.55);font-size:1rem;font-family:Heebo,sans-serif;">'
        f'מערכת ניתוח תוצאות</span>'
        f'</div>'
        f'<div class="nav-links">{links_html}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


_FOOTER = (
    '<div class="footer"><strong>אדמה אפיון ושיקום אתרים בע"מ</strong>'
    ' · מערכת ניתוח תוצאות מעבדה · גרסה 2.0</div>'
)

# ══════════════════════════════════════════════════════════════════
# SIDEBAR — soil page only
# ══════════════════════════════════════════════════════════════════
if page == "soil":
    # Un-hide sidebar for soil page (overrides global "display:none")
    st.markdown(
        '<style>'
        '[data-testid="stAppViewContainer"] > section[data-testid="stSidebar"]{'
        'display:flex!important;'
        'background:#1e293b;'
        'min-width:15rem!important;max-width:16rem!important;width:15.5rem!important;'
        'flex-shrink:0!important;}'
        '[data-testid="stSidebarContent"]{direction:rtl;}'
        '[data-testid="stSidebar"] *{color:#e2e8f0!important;}'
        '[data-testid="stSidebar"] input{color:#111827!important;}'
        '[data-testid="stSidebar"] [data-baseweb="select"] span,'
        '[data-testid="stSidebar"] [data-baseweb="select"] div{color:#111827!important;}'
        '[data-testid="stSidebar"] .stSelectbox label,'
        '[data-testid="stSidebar"] .stTextInput label,'
        '[data-testid="stSidebar"] .stCheckbox label{color:#94a3b8!important;}'
        '[data-testid="stSidebarCollapseButton"],'
        '[data-testid="stSidebarNavCollapseButton"],'
        '[data-testid="collapsedControl"],'
        'button[aria-label="Close sidebar"],'
        'button[aria-label="Open sidebar"],'
        'button[aria-label="פתח סרגל צד"],'
        'button[aria-label="סגור סרגל צד"]{display:none!important;}'
        '</style>',
        unsafe_allow_html=True,
    )

    _nav_logo_small = (
        f'<img src="data:image/png;base64,{LOGO_B64}" style="height:34px;display:block;">'
        if LOGO_B64 else '<span style="font-size:1.3rem;">🧪</span>'
    )

    with st.sidebar:
        st.markdown(
            f'<div style="text-align:center;padding:0.5rem 0 0.75rem;">'
            f'<div style="background:white;border-radius:2px;padding:0.6rem 0.8rem;'
            f'margin-bottom:0.5rem;display:inline-block;width:90%;">'
            f'{_nav_logo_small}</div>'
            f'<div style="font-size:0.7rem;color:#94a3b8;margin-top:4px;">Lab Results Analyzer</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        st.markdown('<hr style="margin:0.5rem 0 1rem;">', unsafe_allow_html=True)

        st.markdown('<div class="sidebar-label">📋 פרטי פרויקט</div>', unsafe_allow_html=True)
        client_name  = st.text_input("שם לקוח",  value="", label_visibility="collapsed",
                                      placeholder="שם לקוח (לדוג׳: סונול)")
        project_name = st.text_input("שם האתר",  value="", label_visibility="collapsed",
                                      placeholder="שם האתר (לדוג׳: צומת שמשון)")

        st.markdown('<div class="sidebar-label">🏭 מעבדה וקטגוריה</div>', unsafe_allow_html=True)
        lab = st.selectbox(
            "מעבדה",
            ["🔍 זיהוי אוטומטי", "KTE", "מכון הנפט", "מכון האנרגיה", "בקטוכם",
             "Alchem", "ALS", "Aminolab", "RJ Lee", "אלכם (XRF)"],
            label_visibility="collapsed",
        )
        category_display = {
            "🔍 זיהוי אוטומטי":           "auto",
            "🪨 קרקע (soil)":             "soil",
            "💧 מי תהום (groundwater)":   "groundwater",
            "🧬 PFAS":                    "pfas",
            "📊 PR format (KTE מתכות)":   "pr",
            "💨 גז קרקע (soil_gas)":      "soil_gas",
            "🪨 גרנולומטריה (grain_size)": "grain_size",
        }
        cat_label    = st.selectbox("קטגוריה", list(category_display.keys()),
                                    label_visibility="collapsed")
        category_raw = category_display[cat_label]

        st.markdown('<hr style="margin:1rem 0 0.5rem;">', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# HOME PAGE
# ══════════════════════════════════════════════════════════════════
if page == "home":
    _render_nav("home")
    st.markdown(
        '<div class="page-wrapper">'
        '<div class="hero">'
        '<h1>מערכת ניתוח נתוני קרקע</h1>'
        '<p>ניתוח אוטומטי של דוחות מעבדה סביבתיים — זיהוי תרכובות, השוואה לערכי סף VSL ו-TIER1, והפקת דוחות Excel מקצועיים</p>'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    with col1:
        if st.button("📊 התחל ניתוח נתונים", use_container_width=True, type="primary"):
            st.query_params["page"] = "soil"
            st.rerun()
    with col2:
        if st.button("💧 דוחות מי תהום", use_container_width=True):
            st.query_params["page"] = "groundwater"
            st.rerun()

    st.markdown(
        '<div class="page-wrapper">'
        '<div class="steps">'
        '<div class="step"><div class="step-num">1</div><div class="step-title">העלאת קבצים</div><div class="step-desc">PDF, Excel או CSV מהמעבדה</div></div>'
        '<div class="step"><div class="step-num">2</div><div class="step-title">עיבוד אוטומטי</div><div class="step-desc">זיהוי מעבדה, תרכובות וערכי סף</div></div>'
        '<div class="step"><div class="step-num">3</div><div class="step-title">הורדת דוח</div><div class="step-desc">Excel מקצועי עם צביעה אוטומטית</div></div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.markdown(_FOOTER, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# GROUNDWATER PAGE
# ══════════════════════════════════════════════════════════════════
elif page == "groundwater":
    _render_nav("groundwater")
    from soil_lab_tool.gw_report_updater import run_update_bytes

    st.markdown(
        '<div class="page-wrapper">'
        '<div class="hero" style="padding:1.5rem 2.5rem;">'
        '<h1 style="font-size:2rem;">💧 עדכון דוח ניטור מי תהום</h1>'
        '<p>העלה את הקבצים — המערכת תוסיף את הדיגום החדש לדוח ולגרפים אוטומטית</p>'
        '</div>',
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    with col1:
        word_file  = st.file_uploader("📄 דוח Word קודם (.docx)", type=["docx"], key="gw_word")
        lab_file   = st.file_uploader("🧪 תוצאות מעבדה — בקטוכם (.pdf)", type=["pdf"], key="gw_lab")
    with col2:
        mk_file    = st.file_uploader("📊 Mann-Kendall (.xls)", type=["xls"], key="gw_mk")
        field_file = st.file_uploader("📋 טופס ממצאי שדה (.pdf, אופציונלי)", type=["pdf"], key="gw_field")

    if word_file and lab_file and mk_file:
        if st.button("⚡ עדכן דוח", type="primary", use_container_width=True):
            with st.spinner("מעבד... ⏳"):
                try:
                    out_word, out_mk = run_update_bytes(
                        word_file.read(),
                        lab_file.read(),
                        mk_file.read(),
                        field_pdf_bytes=field_file.read() if field_file else None,
                    )
                    st.success("✅ הדוח עודכן בהצלחה!")
                    c1, c2 = st.columns(2)
                    with c1:
                        st.download_button(
                            "⬇️ הורד דוח Word מעודכן",
                            data=out_word,
                            file_name="דוח_ניטור_מעודכן.docx",
                            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                            use_container_width=True,
                        )
                    with c2:
                        st.download_button(
                            "⬇️ הורד Mann-Kendall מעודכן",
                            data=out_mk,
                            file_name="mann_kendall_מעודכן.xls",
                            mime="application/vnd.ms-excel",
                            use_container_width=True,
                        )
                except Exception as e:
                    st.error(f"שגיאה בעיבוד: {e}")
    else:
        st.markdown(
            '<div class="card"><div class="card-header">📋 שלבי השימוש</div>'
            '<div class="steps">'
            '<div class="step"><div class="step-num">1</div><div class="step-title">דוח Word קודם</div><div class="step-desc">הדוח מהסבב הקודם (.docx)</div></div>'
            '<div class="step"><div class="step-num">2</div><div class="step-title">תוצאות מעבדה</div><div class="step-desc">PDF של בקטוכם מהדיגום החדש</div></div>'
            '<div class="step"><div class="step-num">3</div><div class="step-title">Mann-Kendall</div><div class="step-desc">קובץ XLS השמור לאתר זה</div></div>'
            '<div class="step"><div class="step-num">4</div><div class="step-title">הורד</div><div class="step-desc">Word + XLS מעודכנים</div></div>'
            '</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown(_FOOTER, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# GUIDE PAGE
# ══════════════════════════════════════════════════════════════════
elif page == "guide":
    _render_nav("guide")
    st.markdown(
        '<div class="page-wrapper">'
        '<div class="hero" style="padding:2rem 2.5rem;">'
        '<h1 style="font-size:2rem;">📖 מדריך למשתמש</h1>'
        '<p>כל מה שצריך לדעת לניתוח תוצאות מעבדה סביבתיות</p>'
        '</div>'

        '<div class="card">'
        '<div class="card-header">🚀 שלבי השימוש במערכת</div>'
        '<div class="steps">'
        '<div class="step">'
        '<div class="step-num">1</div>'
        '<div class="step-title">בחר מעבדה</div>'
        '<div class="step-desc">בחר מתוך הרשימה או השתמש בזיהוי אוטומטי</div>'
        '</div>'
        '<div class="step">'
        '<div class="step-num">2</div>'
        '<div class="step-title">העלה קובץ</div>'
        '<div class="step-desc">PDF, Excel (XLSX/XLS) או CSV</div>'
        '</div>'
        '<div class="step">'
        '<div class="step-num">3</div>'
        '<div class="step-title">בחר ערכי סף</div>'
        '<div class="step-desc">VSL, TIER1, GW לפי סוג האתר</div>'
        '</div>'
        '<div class="step">'
        '<div class="step-num">4</div>'
        '<div class="step-title">הורד דוח</div>'
        '<div class="step-desc">Excel + Word עם צביעה אוטומטית</div>'
        '</div>'
        '</div>'
        '</div>'

        '<div class="card">'
        '<div class="card-header">🏭 מעבדות נתמכות</div>'
        '<table class="lab-table">'
        '<thead><tr>'
        '<th>מעבדה</th>'
        '<th>פורמטים נתמכים</th>'
        '<th>סוגי ניתוח</th>'
        '</tr></thead>'
        '<tbody>'
        '<tr><td><strong>Alchem</strong></td>'
        '<td><span class="badge badge-blue">Excel</span><span class="badge badge-green">PDF</span></td>'
        '<td>קרקע, VOC, SVOC, TPH, מתכות, גז קרקע</td></tr>'
        '<tr><td><strong>KTE</strong></td>'
        '<td><span class="badge badge-blue">Excel</span><span class="badge badge-gray">XML</span></td>'
        '<td>קרקע, מי תהום, גז קרקע, PFAS</td></tr>'
        '<tr><td><strong>מכון הנפט</strong></td>'
        '<td><span class="badge badge-blue">Excel</span></td>'
        '<td>קרקע (VOC, SVOC, מתכות)</td></tr>'
        '<tr><td><strong>מכון האנרגיה</strong></td>'
        '<td><span class="badge badge-blue">Excel</span></td>'
        '<td>קרקע, גז קרקע</td></tr>'
        '<tr><td><strong>בקטוכם</strong></td>'
        '<td><span class="badge badge-green">PDF</span><span class="badge badge-gray">CSV</span></td>'
        '<td>קרקע (SVOC, ICP, TPH), מי תהום</td></tr>'
        '<tr><td><strong>ALS</strong></td>'
        '<td><span class="badge badge-blue">Excel</span><span class="badge badge-green">PDF</span></td>'
        '<td>קרקע, מי תהום, גרנולומטריה</td></tr>'
        '<tr><td><strong>Aminolab</strong></td>'
        '<td><span class="badge badge-green">PDF</span></td>'
        '<td>מי תהום</td></tr>'
        '<tr><td><strong>RJ Lee</strong></td>'
        '<td><span class="badge badge-blue">Excel</span></td>'
        '<td>PFAS (Method 1633)</td></tr>'
        '<tr><td><strong>XRF (אלכם)</strong></td>'
        '<td><span class="badge badge-blue">Excel</span><span class="badge badge-gray">CSV</span></td>'
        '<td>קרקע — ניתוח מתכות XRF</td></tr>'
        '</tbody></table>'
        '</div>'

        '<div class="card">'
        '<div class="card-header">מקרא צבעים בדוח הפלט</div>'
        '<table class="lab-table">'
        '<thead><tr><th>צבע</th><th>משמעות</th></tr></thead>'
        '<tbody>'
        '<tr><td><span style="background:#FFFF00;padding:4px 20px;border-radius:4px;border:1px solid #ccc;">צהוב</span></td>'
        '<td>חריגה מ-VSL</td></tr>'
        '<tr><td><span style="background:#ADD8E6;padding:4px 20px;border-radius:4px;border:1px solid #ccc;">כחול</span></td>'
        '<td>חריגה מ-TIER1 מגורים</td></tr>'
        '<tr><td><span style="background:#FFB6C1;padding:4px 20px;border-radius:4px;border:1px solid #ccc;">ורוד</span></td>'
        '<td>חריגה מ-TIER1 תעשייה</td></tr>'
        '<tr><td><span style="background:#D3D3D3;padding:4px 20px;border-radius:4px;border:1px solid #ccc;">אפור</span></td>'
        '<td>LOQ גבוה מסף</td></tr>'
        '</tbody></table>'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    st.markdown(_FOOTER, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# SOIL PAGE — existing analysis logic
# ══════════════════════════════════════════════════════════════════
elif page == "soil":
    # nav bar
    _render_nav("soil")

    # ── UPLOAD ────────────────────────────────────────────────────
    st.markdown('<div class="section-card">', unsafe_allow_html=True)

    col_up, col_meta = st.columns([3, 1])

    with col_up:
        uploaded_files = st.file_uploader(
            "גרור קבצים לכאן או לחץ לבחירה",
            type=["xlsx", "xls", "csv", "pdf"],
            accept_multiple_files=True,
            help="ניתן להעלות מספר קבצים מאותה מעבדה | XLSX / XLS / CSV / PDF",
            label_visibility="visible",
        )

    with col_meta:
        _display_lab = lab
        if uploaded_files and lab == "🔍 זיהוי אוטומטי":
            try:
                _peek = uploaded_files[0].getvalue()
                _det  = auto_detect_lab(uploaded_files[0].name, _peek)
                _display_lab = _det or "KTE"
            except Exception:
                _display_lab = "?"
        elif lab == "🔍 זיהוי אוטומטי":
            _display_lab = "—"

        cat_clean = cat_label.split(" ", 1)[-1] if " " in cat_label else cat_label
        st.markdown(f"""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
                    padding:0.9rem 1rem;margin-top:1.75rem;text-align:center;">
          <div style="font-size:0.65rem;color:#94a3b8;font-weight:700;letter-spacing:0.8px;
                      text-transform:uppercase;margin-bottom:4px;">מעבדה</div>
          <div style="font-size:1.2rem;font-weight:800;color:#1e3a4f;">{_display_lab}</div>
          <div style="font-size:0.7rem;color:#64748b;margin-top:4px;">{cat_clean}</div>
        </div>
        """, unsafe_allow_html=True)

    pid_upload_file = st.file_uploader(
        "העלאת ממצאי שדה (PID) — אופציונלי",
        type=["xlsx"],
        key="pid_upload",
        help="עמודה 0 = שם קידוח (מולא רק בשורה הראשונה), עמודה 7 = PID [ppm]",
    )
    if pid_upload_file is not None:
        try:
            _pm = _parse_pid_file(pid_upload_file)
            st.session_state["pid_map"] = _pm
            st.caption(f"✅ נתוני PID נטענו — {len(_pm)} קידוחים")
        except Exception as _pid_err:
            st.warning(f"⚠️ שגיאת קריאת קובץ PID: {_pid_err}")
            st.session_state["pid_map"] = {}
    else:
        st.session_state["pid_map"] = {}

    st.markdown('</div>', unsafe_allow_html=True)

    if not uploaded_files:
        st.markdown("""
        <div class="info-banner" style="margin:0 2rem;">
          ℹ️ העלה קובץ דוח מעבדה כדי להתחיל — המערכת תזהה אוטומטית את סוג הניתוח
        </div>
        """, unsafe_allow_html=True)
        st.markdown(_FOOTER, unsafe_allow_html=True)
        st.stop()

    # ── PARSE ─────────────────────────────────────────────────────
    all_raw: list[tuple[str, bytes]] = [(uf.name, uf.read()) for uf in uploaded_files]
    fname     = " | ".join(f for f, _ in all_raw)
    raw_bytes = all_raw[0][1]

    if lab == "🔍 זיהוי אוטומטי":
        detected_lab = None
        for _fn, _fb in all_raw:
            _det = auto_detect_lab(_fn, _fb)
            if _det and _det != "KTE":
                detected_lab = _det
                break
        if not detected_lab:
            detected_lab = auto_detect_lab(all_raw[0][0], all_raw[0][1]) or "KTE"
        lab = detected_lab

    if category_raw == 'auto':
        category = None
        for _fn, _fb in all_raw:
            _cat = auto_detect_category(_fn, _fb, lab=lab)
            if _cat and _cat != "soil":
                category = _cat
                break
        if not category:
            category = auto_detect_category(all_raw[0][0], all_raw[0][1], lab=lab) or "soil"
        cat_info = f"זוהה אוטומטית: **{category}**"
    else:
        category = category_raw
        cat_info  = f"קטגוריה: **{category}**"

    try:
        try:
            parser = get_parser(lab, category)
        except KeyError:
            fallback = "soil"
            st.warning(f"⚠️ אין parser עבור {lab} / {category}. מנסה: **{fallback}**")
            category = fallback
            cat_info  = f"ברירת מחדל: **{fallback}**"
            parser   = get_parser(lab, fallback)
    except Exception as e:
        st.error(f"שגיאת טעינת parser: {e}")
        st.exception(e)
        st.stop()

    all_records:    list[dict] = []
    file_summaries: list[dict] = []
    n_files = len(all_raw)

    _pdf_raws    = [b for n, b in all_raw if n.lower().endswith(".pdf")]
    _data_files  = [(n, b) for n, b in all_raw if not n.lower().endswith(".pdf")]
    _parse_files = _data_files if _data_files else all_raw

    with st.spinner(f"מנתח {'קבצים' if n_files > 1 else 'קובץ'}..."):
        for fname_i, raw_i in _parse_files:
            try:
                try:
                    file_records = parser.parse(io.BytesIO(raw_i), pdf_bytes=_pdf_raws)
                except TypeError:
                    file_records = parser.parse(io.BytesIO(raw_i))
                all_records.extend(file_records)
                file_summaries.append({"name": fname_i, "records": len(file_records), "ok": True})
            except Exception as e:
                st.error(f"שגיאת פרסינג: {fname_i} — {e}")
                file_summaries.append({"name": fname_i, "records": 0, "ok": False})

    records = all_records

    if not records:
        st.markdown(
            '<div class="info-banner" style="margin:0 2rem;">'
            '⚠️ לא נמצאו רשומות — בדוק פורמט הקובץ ובחירת מעבדה / קטגוריה'
            '</div>',
            unsafe_allow_html=True,
        )
        st.stop()

    # ── stats ─────────────────────────────────────────────────────
    by_type  = collections.Counter(
        r.get('analysis_type', '?') for r in records
        if r.get('value') is not None and r.get('flag') not in ('<LOQ', 'ND')
    )
    samples  = sorted(set(r['sample_id'] for r in records))
    detected = [r for r in records if r.get('flag') not in ('ND', '<LOD') and r.get('value') is not None]

    st.markdown(f"""
    <div class="success-banner" style="margin:0 2rem 0.5rem;">
      ✅ {cat_info} &nbsp;|&nbsp; Parser: <code>{type(parser).__name__}</code>
      {"&nbsp;|&nbsp; " + " ".join(f'<b>{s["name"]}</b>: {s["records"]} רשומות' for s in file_summaries) if n_files > 1 else ""}
    </div>
    """, unsafe_allow_html=True)

    _mc1, _mc2, _mc3, _mc4 = st.columns(4)
    with _mc1: st.metric("סה\"כ רשומות",  f"{len(records):,}")
    with _mc2: st.metric("ערכים מזוהים",  f"{len(detected):,}")
    with _mc3: st.metric("דגימות",        f"{len(samples):,}")
    with _mc4: st.metric("סוגי ניתוח",   f"{len(by_type):,}")

    BADGE_COLORS = {
        "SOIL_GAS_VOC": "#7c3aed", "SOIL_VOC":    "#0d9488",
        "SOIL_TPH":     "#0891b2", "SOIL_MBTEX":  "#0f766e",
        "SOIL_METALS":  "#4f46e5", "SOIL_PFAS":   "#db2777",
        "GW_VOC":       "#4a7a8a", "GW_PFAS":     "#9333ea",
        "LOWFLOW":      "#6b7280",
    }
    _TYPE_LABELS = {
        "SOIL_VOC":     "VOC",
        "SOIL_SVOC":    "SVOC",
        "SOIL_TPH":     "TPH",
        "SOIL_PFAS":    "PFAS",
        "SOIL_METALS":  "מתכות",
        "GW_VOC":       "מי תהום VOC",
        "SOIL_GAS_VOC": "גז קרקע",
        "GAS_VOC":      "גז קרקע",
    }
    badges = " ".join(
        f'<span class="type-badge" style="background:{BADGE_COLORS.get(t,"#94a3b8")};">'
        f'{_TYPE_LABELS.get(t, t)}</span>'
        for t in by_type
    )
    st.markdown(f'<div style="margin:0.5rem 2rem;">{badges}</div>', unsafe_allow_html=True)

    # ── THRESHOLD SELECTION ───────────────────────────────────────
    found_atypes  = list(by_type.keys())
    has_soil      = (any(t in found_atypes for t in ("SOIL_VOC","SOIL_TPH","SOIL_METALS","SOIL_MBTEX","SOIL_SVOC","XRF"))
                     or category in ("soil", "soil_pdf"))
    has_soil_pfas = "SOIL_PFAS" in found_atypes
    has_soil_gas  = "SOIL_GAS_VOC" in found_atypes
    has_gw        = any(t.startswith("GW_") for t in found_atypes)

    selected_thresholds: list[str] = []

    _SENS_MAP  = {"רגיש מאוד": "vh", "רגיש/בינוני": "hm", "לא רגיש": "low", "—": None}
    _DEPTH_MAP = {"0-6מ'": "0_6", ">6מ'": "6"}

    def _soil_tier1_key(land_use: str, sens_code, depth_label) -> str | None:
        if not sens_code:
            return None
        pfx = "RES" if land_use == "res" else "IND"
        if sens_code == "vh":  return f"TIER1_{pfx}_SOIL_VH"
        if sens_code == "hm":
            d = _DEPTH_MAP.get(depth_label, "0_6")
            return f"TIER1_{pfx}_SOIL_HM_{d}"
        if sens_code == "low": return f"TIER1_{pfx}_SOIL_LOW"
        return None

    _PFAS_SENS_MAP = {"רגישות גבוהה מאוד": "vh", "רגישות גבוהה/בינונית": "hm", "רגישות נמוכה": "low", "—": None}

    def _pfas_tier1_key(land_use: str, sens_code, depth_label) -> str | None:
        if not sens_code:
            return None
        pfx = "RES" if land_use == "res" else "IND"
        if sens_code == "vh":  return f"PFAS_TIER1_{pfx}_VERY_HIGH"
        if sens_code == "hm":
            return f"PFAS_TIER1_{pfx}_0_6" if depth_label == "0-6מ'" else f"PFAS_TIER1_{pfx}_6PLUS"
        if sens_code == "low": return f"PFAS_TIER1_{pfx}_NO_GW"
        return None

    any_shown = False

    st.markdown('<div class="section-card">', unsafe_allow_html=True)

    if has_soil:
        any_shown = True
        st.markdown("##### 🪨 קרקע")
        st.markdown('<span class="thresh-pill-marker"></span>', unsafe_allow_html=True)
        tog1, tog2, _tog_rest = st.columns([1, 1, 4])
        with tog1:
            use_vsl   = st.checkbox("📊 VSL",   value=True, key="vsl_cb")
        with tog2:
            use_tier1 = st.checkbox("🏗️ TIER1", value=True, key="tier1_cb")

        if use_tier1:
            sub1, sub2 = st.columns(2)
            with sub1:
                land_use_sel = st.radio(
                    "שימוש קרקע", ["תעשייה", "מגורים"],
                    horizontal=True, key="tier1_land_use",
                    label_visibility="collapsed",
                )
            with sub2:
                tier1_sens_sel = st.selectbox(
                    "רגישות", [
                        "רגיש מאוד",
                        "רגיש/בינוני 0-6מ'",
                        "רגיש/בינוני >6מ'",
                        "לא רגיש",
                    ],
                    key="tier1_sens", label_visibility="collapsed",
                )

        if use_vsl:
            selected_thresholds.append("VSL_SOIL")
        if use_tier1:
            _pfx = "RES" if land_use_sel == "מגורים" else "IND"
            _TIER1_SOIL_MAP = {
                "רגיש מאוד":          f"TIER1_{_pfx}_SOIL_VH",
                "רגיש/בינוני 0-6מ'": f"TIER1_{_pfx}_SOIL_HM_0_6",
                "רגיש/בינוני >6מ'":  f"TIER1_{_pfx}_SOIL_HM_6",
                "לא רגיש":            f"TIER1_{_pfx}_SOIL_LOW",
            }
            k = _TIER1_SOIL_MAP.get(tier1_sens_sel)
            if k:
                selected_thresholds.append(k)

    if has_soil_pfas:
        any_shown = True
        st.markdown("##### 🧬 קרקע PFAS")
        cp_vsl, cp_res, cp_ind = st.columns(3)

        with cp_vsl:
            st.markdown('<div style="font-size:0.85rem;font-weight:700;color:#374151;margin-bottom:6px;">VSL — ישיר</div>', unsafe_allow_html=True)
            use_pfas_vsl = st.checkbox("PFAS VSL", value=True, key="pfas_vsl")

        with cp_res:
            st.markdown('<div style="font-size:0.85rem;font-weight:700;color:#374151;margin-bottom:6px;">Tier 1 מגורים (Residential)</div>', unsafe_allow_html=True)
            pfas_sens_res = st.selectbox("רגישות", ["—", "רגישות גבוהה מאוד", "רגישות גבוהה/בינונית", "רגישות נמוכה"],
                                         key="pfas_sens_res", label_visibility="collapsed")
            pfas_depth_res = None
            if pfas_sens_res == "רגישות גבוהה/בינונית":
                pfas_depth_res = st.radio("עומק", ["0-6מ'", ">6מ'"], horizontal=True,
                                          key="pfas_depth_res", label_visibility="collapsed")

        with cp_ind:
            st.markdown('<div style="font-size:0.85rem;font-weight:700;color:#374151;margin-bottom:6px;">Tier 1 תעשייה (Industrial)</div>', unsafe_allow_html=True)
            pfas_sens_ind = st.selectbox("רגישות", ["—", "רגישות גבוהה מאוד", "רגישות גבוהה/בינונית", "רגישות נמוכה"],
                                         key="pfas_sens_ind", label_visibility="collapsed")
            pfas_depth_ind = None
            if pfas_sens_ind == "רגישות גבוהה/בינונית":
                pfas_depth_ind = st.radio("עומק", ["0-6מ'", ">6מ'"], horizontal=True,
                                          key="pfas_depth_ind", label_visibility="collapsed")

        if use_pfas_vsl:
            selected_thresholds.append("PFAS_VSL")
        k = _pfas_tier1_key("res", _PFAS_SENS_MAP.get(pfas_sens_res), pfas_depth_res)
        if k: selected_thresholds.append(k)
        k = _pfas_tier1_key("ind", _PFAS_SENS_MAP.get(pfas_sens_ind), pfas_depth_ind)
        if k: selected_thresholds.append(k)

    if has_soil_gas:
        any_shown = True
        st.markdown("##### 💨 גז קרקע VOC")
        gas_indoor_type = st.radio(
            "סוג ערך סף — אוויר פנים מבני:",
            options=[
                "גז קרקע — הגנה על אוויר פנים מבני",
                "אוויר תוך מבני (Ambient Air)",
            ],
            index=0,
            key="gas_indoor_type",
            horizontal=True,
        )
        sg_col_r, sg_col_i = st.columns(2)
        with sg_col_r:
            st.markdown('<div style="font-size:0.8rem;font-weight:600;color:#374151;">Tier 1 מגורים</div>', unsafe_allow_html=True)
            sg_res_in  = st.checkbox("Indoor — פנים",  value=True,  key="sg_res_in")
            sg_res_out = st.checkbox("Outdoor — חוץ",  value=False, key="sg_res_out")
        with sg_col_i:
            st.markdown('<div style="font-size:0.8rem;font-weight:600;color:#374151;">Tier 1 תעשייה</div>', unsafe_allow_html=True)
            sg_ind_in  = st.checkbox("Indoor — פנים",  value=False, key="sg_ind_in")
            sg_ind_out = st.checkbox("Outdoor — חוץ",  value=False, key="sg_ind_out")
        _use_ambient = gas_indoor_type == "אוויר תוך מבני (Ambient Air)"
        if sg_res_in:  selected_thresholds.append("GAS_AMBIENT_RES" if _use_ambient else "GAS_INDOOR_RES")
        if sg_res_out: selected_thresholds.append("GAS_OUTDOOR_RES")
        if sg_ind_in:  selected_thresholds.append("GAS_AMBIENT_IND" if _use_ambient else "GAS_INDOOR_IND")
        if sg_ind_out: selected_thresholds.append("GAS_OUTDOOR_IND")

    if has_gw:
        any_shown = True
        st.markdown("##### 💧 מי תהום")
        use_gw = st.checkbox('ערך סף מי"ת (GW Standard)', value=True, key="gw_cb")
        if use_gw: selected_thresholds.append("GW")

    if not any_shown:
        st.info("ℹ️ LOWFLOW — ממצאי שדה בלבד, ללא ערכי סף")
    elif not selected_thresholds:
        st.warning("⚠️ לא נבחרו ערכי סף — הדוח ייצא ללא עמודות השוואה")

    has_tph_and_voc   = "SOIL_TPH" in found_atypes and "SOIL_VOC"   in found_atypes
    has_tph_and_mbtex = "SOIL_TPH" in found_atypes and "SOIL_MBTEX" in found_atypes
    combine_tph_voc   = False
    combine_tph_mbtex = False

    if has_tph_and_voc or has_tph_and_mbtex:
        st.markdown('<div style="margin-top:0.5rem;"></div>', unsafe_allow_html=True)
        cc1, cc2 = st.columns(2)
        with cc1:
            if has_tph_and_voc:
                combine_tph_voc = st.checkbox("שלב TPH + BTEX בגיליון אחד", value=False, key="combine_tph_voc")
        with cc2:
            if has_tph_and_mbtex:
                combine_tph_mbtex = st.checkbox("שלב TPH + MBTEX בגיליון אחד", value=False, key="combine_tph_mbtex")

    st.markdown('</div>', unsafe_allow_html=True)

    # ── BUILD EXCEL + DOWNLOAD ────────────────────────────────────
    st.markdown('<div class="section-card">', unsafe_allow_html=True)

    _sniff   = raw_bytes.lstrip()[:200]
    _is_kte_gw = (
        lab == "KTE" and category == "groundwater" and
        (b"<?xml" in _sniff or b"<Workbook" in _sniff)
    )

    pid_map = st.session_state.get("pid_map", {})

    excel_buf = io.BytesIO()
    excel_ok  = False
    word_buf  = io.BytesIO()
    word_ok   = False

    if _is_kte_gw:
        try:
            from core.excel_output import build_kte_gw_btex_simple_from_xml
            build_kte_gw_btex_simple_from_xml(raw_bytes, excel_buf)
            excel_ok = True
        except Exception as e:
            st.error(f"שגיאת בניית Excel: {e}")
            st.exception(e)
    elif lab.lower() in ("xrf",) or lab in ("אלכם", "אלכם (XRF)"):
        try:
            build_xrf_simple_excel(
                records,
                excel_buf,
                threshold_manager   = tm,
                selected_thresholds = selected_thresholds,
                project_name        = project_name,
                client              = client_name,
                report_date         = date.today().strftime('%d.%m.%Y'),
            )
            excel_buf.seek(0)
            excel_ok = True
        except Exception as e:
            st.error(f"שגיאת בניית Excel: {e}")
            st.exception(e)
    else:
        thresh_display = ", ".join(tm.threshold_label(k) for k in selected_thresholds) or "ללא ערכי סף"
        st.caption(f"📌 ערכי סף: **{thresh_display}**")
        try:
            builder = LabReportExcel(
                records             = records,
                threshold_manager   = tm,
                output_path         = excel_buf,
                project_name        = project_name,
                client              = client_name,
                report_date         = date.today().strftime('%d.%m.%Y'),
                selected_thresholds = selected_thresholds,
                combine_tph_voc     = combine_tph_voc,
                combine_tph_mbtex   = combine_tph_mbtex,
                pid_map             = pid_map,
            )
            builder.build()
            excel_buf.seek(0)
            excel_ok = True
            try:
                LabReportWord(
                    records             = records,
                    threshold_manager   = tm,
                    output_path         = word_buf,
                    project_name        = project_name,
                    client              = client_name,
                    report_date         = date.today().strftime('%d.%m.%Y'),
                    selected_thresholds = selected_thresholds,
                    combine_tph_voc     = combine_tph_voc,
                    combine_tph_mbtex   = combine_tph_mbtex,
                ).build()
                word_buf.seek(0)
                word_ok = True
            except Exception as e:
                st.warning(f"⚠️ שגיאת בניית Word: {e}")
        except Exception as e:
            st.error(f"שגיאת בניית Excel: {e}")
            st.exception(e)

    if excel_ok:
        def _safe(s: str) -> str:
            import re as _re
            return _re.sub(r'[\\/*?:"<>|\s]+', '_', s.strip()).strip('_') or 'x'
        _parts = ["lab_report"]
        if client_name.strip():  _parts.append(_safe(client_name))
        if project_name.strip(): _parts.append(_safe(project_name))
        out_filename = f"{'_'.join(_parts)}.xlsx"
        size_kb = len(excel_buf.getvalue()) / 1024

        dl_col, prev_col, info_col = st.columns([2, 1.5, 1])
        with dl_col:
            st.download_button(
                label     = "⬇️ הורד דוח Excel",
                data      = excel_buf.getvalue(),
                file_name = out_filename,
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with prev_col:
            if st.button("👁️ תצוגה מקדימה", use_container_width=True, key="xl_preview_btn"):
                _preview_dialog(records, tm, project_name, client_name,
                                date.today().strftime('%d.%m.%Y'),
                                selected_thresholds, combine_tph_voc,
                                combine_tph_mbtex, pid_map)
        with info_col:
            st.markdown(f"""
            <div style="padding:0.5rem 0;font-size:0.82rem;color:#64748b;direction:rtl;">
              <div>📄 <b>{out_filename}</b></div>
              <div>📦 גודל: {size_kb:.1f} KB</div>
              <div>📅 {date.today().strftime('%d.%m.%Y')}</div>
            </div>
            """, unsafe_allow_html=True)

        if word_ok:
            word_out = out_filename.replace(".xlsx", ".docx")
            st.download_button(
                label     = "⬇️ הורד דוח Word",
                data      = word_buf.getvalue(),
                file_name = word_out,
                mime      = "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                use_container_width=False,
                key       = "word_dl_btn",
            )

    st.markdown('</div>', unsafe_allow_html=True)

    # ── file info footer ──────────────────────────────────────────
    st.markdown(
        f'<div style="text-align:center;color:#94a3b8;font-size:0.75rem;margin:0.5rem 2rem 0;">'
        f'🔬 {lab} / {category} &nbsp;·&nbsp; '
        f'📁 {fname[:80]}{"…" if len(fname)>80 else ""} &nbsp;·&nbsp; '
        f'📅 {date.today().strftime("%d.%m.%Y")}'
        f'</div>',
        unsafe_allow_html=True,
    )

    st.markdown(_FOOTER, unsafe_allow_html=True)
