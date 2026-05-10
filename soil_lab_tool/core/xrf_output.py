"""
core/xrf_output.py
-------------------
Excel report for XRF soil-metals data — landscape layout matching LabReportExcel.

Layout (Index × Elements — same structure as _write_landscape in excel_output.py):
  Row 1  : Header — שם פרויקט: | תאריך: | מזמין:  (3 merged cells)
  Row 2  : Column headers — שם קידוח | עומק [מ'] | PID [ppm] | Mo | Zr | Pb | …
  Row 3  : CAS numbers
  Row 4  : Units (mg/kg for all elements)
  Row 5+ : Threshold rows (one per selected key; before sample data)
  Row N+ : Data rows — Index value | (empty) | (empty) | measurements
  Legend + source footnotes
"""
from __future__ import annotations

import io
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from core.threshold_manager import ThresholdManager

# ── Style constants — identical to excel_output.py ───────────────────────────
ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
GRAY   = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
WRAP_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

FHE = {"name": "David",           "size": 9}
FEN = {"name": "Times New Roman", "size": 8}

_UNDEF_FONT = Font(name="David", size=9, color="808080", italic=True)

_THRESHOLD_SOURCES: dict[str, str] = {
    "VSL_SOIL":              "Soil VSL, Rev. 7, 12/24",
    "TIER1_RES_SOIL_VH":     "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_HM_0_6": "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_HM_6":   "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_LOW":    "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_IND_SOIL_VH":     "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_HM_0_6": "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_HM_6":   "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_LOW":    "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "GW":                    "Groundwater Standard, Rev.7, 12/24",
}


def _font(val, bold: bool = False) -> Font:
    """David 9 for Hebrew/mixed/numbers; Times New Roman 8 for pure English."""
    s = str(val) if val is not None else ""
    has_hebrew  = any('א' <= c <= 'ת' for c in s)
    has_english = any(c.isalpha() and c.isascii() for c in s)
    base = FEN if (has_english and not has_hebrew) else FHE
    return Font(**base, bold=bold)


def _round_thresh(v):
    if v is None or not isinstance(v, (int, float)):
        return v
    return round(v, 2)


def _fmt_value(value, flag: str, lod) -> object:
    """Display value: number, '< {lod}', or '< LOD'."""
    if flag == "" and value is not None:
        v = float(value)
        if v == int(v) and abs(v) < 1e9:
            return int(v)
        return round(v, 3)
    if flag == "<LOD" and lod is not None:
        lf = float(lod)
        s  = f"{lf:.4f}".rstrip("0").rstrip(".")
        return f"< {s}"
    return "< LOD"


def _strictest(t_vals: dict):
    vals = [v for v in t_vals.values() if v is not None]
    return min(vals) if vals else None


def _source_notes(thresh_keys: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for k in thresh_keys:
        src = _THRESHOLD_SOURCES.get(k)
        if src and src not in seen:
            seen.add(src)
            out.append(src)
    return out


# ── Main builder ──────────────────────────────────────────────────────────────

def build_xrf_simple_excel(
    records: list[dict],
    output_path: str | io.BytesIO,
    threshold_manager: "ThresholdManager | None" = None,
    selected_thresholds: list[str] | None = None,
    project_name: str = "",
    client: str = "",
    report_date: str = "",
) -> None:
    """
    Write a professional XRF Excel report matching the LabReportExcel landscape layout.

    Parameters
    ----------
    records             : parsed XRF records (sample_id=Index, compound=symbol)
    output_path         : file path or BytesIO
    threshold_manager   : loaded ThresholdManager (optional)
    selected_thresholds : list of threshold keys (e.g. ['VSL_SOIL', 'TIER1_RES_SOIL_VH'])
    project_name / client / report_date : shown in header row
    """
    from core.threshold_manager import THRESHOLD_LABELS

    thresh_keys = selected_thresholds or []
    has_thresh  = bool(threshold_manager and thresh_keys)

    # ── Collect ordered unique indices and element symbols ────────────
    seen_idx: set  = set()
    indices:  list = []
    seen_sym: set  = set()
    symbols:  list = []
    cas_map:  dict = {}   # symbol → CAS

    for r in records:
        idx = str(r.get("sample_id", "")).strip()
        sym = r.get("compound", "")
        cas = r.get("cas", "")
        if idx and idx not in seen_idx:
            seen_idx.add(idx)
            indices.append(idx)
        if sym and sym not in seen_sym:
            seen_sym.add(sym)
            symbols.append(sym)
            if cas:
                cas_map[sym] = cas

    # Sort indices numerically when possible (76, 100, 124, 163 …)
    def _sort_key(s: str):
        try:
            return (0, float(s))
        except ValueError:
            return (1, s)
    indices.sort(key=_sort_key)

    # lookup: index → symbol → (raw_value, flag, lod)
    # First reading for each sample/element wins — preserves original file values.
    lookup: dict[str, dict[str, tuple]] = {}
    for r in records:
        idx  = str(r.get("sample_id", "")).strip()
        sym  = r.get("compound", "")
        val  = r.get("value")
        flag = r.get("flag", "ND")
        lod  = r.get("lod")
        inner = lookup.setdefault(idx, {})
        if sym not in inner:   # keep only the first reading per sample/element
            inner[sym] = (val, flag, lod)

    # Pre-compute threshold values: thresh_key → symbol → numeric value or None
    thresh_vals: dict[str, dict[str, object]] = {}
    if has_thresh:
        for key in thresh_keys:
            thresh_vals[key] = {}
            for sym in symbols:
                cas = cas_map.get(sym, "")
                tv  = threshold_manager.get_threshold_with_name(cas, key, compound_name=sym)
                thresh_vals[key][sym] = tv

    # ── Workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "XRF מתכות"
    ws.sheet_view.rightToLeft = True

    # Fixed columns: שם קידוח | עומק [מ'] | PID [ppm]
    N_FIXED       = 3
    total_cols    = N_FIXED + len(symbols)
    cmp_col_start = N_FIXED + 1

    # ── Row 1: header (שם פרויקט | תאריך | מזמין) ────────────────────
    parts = [
        ("שם פרויקט:", project_name),
        ("תאריך:",     report_date),
        ("מזמין:",     client),
    ]
    span = max(1, total_cols // len(parts))
    for i, (label, val) in enumerate(parts):
        col_start = i * span + 1
        col_end   = (i + 1) * span if i < len(parts) - 1 else total_cols
        ws.merge_cells(start_row=1, start_column=col_start,
                       end_row=1, end_column=col_end)
        c = ws.cell(row=1, column=col_start, value=f"{label}  {val}")
        c.font      = Font(**FHE, bold=True)
        c.alignment = WRAP_C
        c.border    = THIN
    ws.row_dimensions[1].height = 20

    # ── Row 2: column headers ─────────────────────────────────────────
    row2 = ["שם קידוח", "עומק [מ']", "PID [ppm]"] + symbols
    for ci, v in enumerate(row2, 1):
        c = ws.cell(row=2, column=ci, value=v)
        c.font      = _font(v, bold=True)
        c.alignment = WRAP_C
        c.border    = THIN
    ws.row_dimensions[2].height = 28

    # ── Row 3: CAS numbers ────────────────────────────────────────────
    row3 = ["CAS Number", "", ""] + [cas_map.get(sym, "") for sym in symbols]
    for ci, v in enumerate(row3, 1):
        c = ws.cell(row=3, column=ci, value=v)
        c.font      = _font(v, bold=(ci == 1))
        c.alignment = WRAP_C
        c.border    = THIN
    ws.row_dimensions[3].height = 20

    # ── Row 4: units ──────────────────────────────────────────────────
    row4 = ["יחידות", "", ""] + ["mg/kg"] * len(symbols)
    for ci, v in enumerate(row4, 1):
        c = ws.cell(row=4, column=ci, value=v)
        c.font      = _font(v, bold=(ci == 1))
        c.alignment = WRAP_C
        c.border    = THIN
    ws.row_dimensions[4].height = 20

    # ── Threshold rows (before sample data, matching landscape layout) ──
    data_row = 5
    if has_thresh:
        for tk in thresh_keys:
            label    = THRESHOLD_LABELS.get(tk, tk)
            lbl_cell = ws.cell(row=data_row, column=1, value=label)
            lbl_cell.font      = _font(label, bold=False)
            lbl_cell.border    = THIN
            lbl_cell.alignment = Alignment(horizontal="right", vertical="center",
                                           wrap_text=True, readingOrder=2)
            for fc in range(2, cmp_col_start):
                ws.cell(row=data_row, column=fc).border = THIN
            for ci, sym in enumerate(symbols, cmp_col_start):
                cas  = cas_map.get(sym, "")
                tval = _round_thresh(
                    threshold_manager.get_threshold_with_name(cas, tk, compound_name=sym)
                )
                c = ws.cell(row=data_row, column=ci)
                c.border = THIN
                if tval is None:
                    c.value     = "לא קיים"
                    c.font      = _UNDEF_FONT
                    c.alignment = CENTER
                else:
                    c.value         = tval
                    c.font          = Font(**FHE)
                    c.number_format = "#,##0.##"
                    c.alignment     = CENTER
            ws.row_dimensions[data_row].height = 20
            data_row += 1

    # ── Sample data rows ──────────────────────────────────────────────
    has_gray = False

    for idx in indices:
        row_data = lookup.get(idx, {})

        # Column 1: Index value (= "שם קידוח")
        c1 = ws.cell(row=data_row, column=1, value=idx)
        c1.font      = _font(idx, bold=True)
        c1.alignment = CENTER
        c1.border    = THIN

        # Columns 2-3: empty depth / PID
        for fc in range(2, cmp_col_start):
            ws.cell(row=data_row, column=fc).border = THIN

        # Element columns
        for ci, sym in enumerate(symbols, cmp_col_start):
            entry = row_data.get(sym)
            if entry is None:
                ws.cell(row=data_row, column=ci).border = THIN
                continue

            raw_val, flag, lod = entry
            display = _fmt_value(raw_val, flag, lod)

            c = ws.cell(row=data_row, column=ci, value=display)
            c.font      = _font(display)
            c.alignment = CENTER
            c.border    = THIN
            if isinstance(display, (int, float)):
                c.number_format = "#,##0.###"

            # Colour coding — same logic as _write_landscape
            if has_thresh:
                t_vals       = {k: thresh_vals.get(k, {}).get(sym) for k in thresh_keys}
                thresh_limit = _strictest(t_vals)
                if thresh_limit is not None:
                    if flag == "" and raw_val is not None:
                        try:
                            if float(raw_val) > thresh_limit:
                                c.fill = ORANGE
                                c.font = Font(**FHE, bold=True)
                        except (TypeError, ValueError):
                            pass
                    elif flag in ("<LOD", "<LOQ", "ND", "<") and lod is not None:
                        try:
                            if float(lod) > thresh_limit:
                                c.fill = GRAY
                                c.font = _font(display, bold=True)
                                has_gray = True
                        except (TypeError, ValueError):
                            pass

        ws.row_dimensions[data_row].height = 16
        data_row += 1

    # ── Legend ────────────────────────────────────────────────────────
    legend_items = [("חריגה מערך סף", ORANGE)]
    if has_gray:
        legend_items.append(("ערך הסף גדול מסף הגילוי", GRAY))
    for i, (label, fill) in enumerate(legend_items):
        c = ws.cell(row=data_row + 1 + i, column=1, value=label)
        c.font      = Font(name="David", size=9, bold=True)
        c.fill      = fill
        c.border    = THIN
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Source footnotes ──────────────────────────────────────────────
    note_row = data_row + 1 + len(legend_items) + 1
    active_keys = [k for k in thresh_keys
                   if any(thresh_vals.get(k, {}).get(sym) is not None for sym in symbols)]
    for note in _source_notes(active_keys):
        ws.cell(row=note_row, column=1, value=f"* {note}").font = Font(
            **FEN, italic=True, color="808080")
        note_row += 1

    # ── Column widths (same as _auto_width) ───────────────────────────
    ws.column_dimensions["A"].width = 13   # שם קידוח / Index
    ws.column_dimensions["B"].width = 11   # עומק [מ']
    ws.column_dimensions["C"].width = 11   # PID [ppm]
    for ci in range(4, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9

    # ── Freeze panes: header rows + fixed cols ────────────────────────
    freeze_row = data_row - len(indices)   # first sample row
    ws.freeze_panes = ws.cell(row=freeze_row, column=cmp_col_start)

    wb.save(output_path)
