"""
core/excel_output.py
---------------------
Builds a multi-sheet, multi-threshold Hebrew RTL Excel report.

One sheet is created per analysis type found in the records.

Portrait layout (compounds as rows):
  A: compound name | B: CAS | C+: threshold cols | next: יחידות | then: sample cols

Landscape layout (samples as rows — soil gas, many-samples):
  Row headers: compound / CAS / unit | Columns: samples
  Threshold rows appended at bottom.

Colour coding:
  Pink    — measured value exceeds Tier 1 Industrial threshold
  Blue    — measured value exceeds Tier 1 Residential threshold
  Orange  — measured value exceeds other threshold (GW / no IND/RES key)
  Yellow  — measured value exceeds VSL (but not Tier 1)
  Gray    — below detection limit, but LOD > threshold (uncertain)

Orientation (per sheet):
  Portrait  (compounds as rows)  — when n_compounds >= n_samples
  Landscape (samples as rows)    — when n_samples > n_compounds

Sheet config:
  SOIL_GAS_VOC → "גז קרקע VOC"   µg/m³
  SOIL_VOC     → "קרקע VOC BTEX" mg/kg
  SOIL_TPH     → "קרקע TPH"      mg/kg
  SOIL_METALS  → "קרקע מתכות"    mg/kg DW
  SOIL_PFAS    → "קרקע PFAS"     ng/g
  GW_VOC       → "מי תהום BTEX"  mg/L
  GW_PFAS      → "מי תהום PFAS"  ng/L
  LOWFLOW      → "pH"             — field parameters, no thresholds
"""

from __future__ import annotations

import copy
import io
import math
import os
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from core.threshold_manager import ThresholdManager, ANALYSIS_THRESHOLDS, THRESHOLD_LABELS
from core.cas_lookup import name_to_cas as _name_to_cas, fuzzy_name_to_cas as _fuzzy_name_to_cas


def _is_ind_key(k: str) -> bool:
    """True for Tier1 Industrial threshold keys (e.g. TIER1_IND_*, TIER1_INDOOR_IND, PFAS_TIER1_IND*).
    Also covers gas thresholds (GAS_AMBIENT_IND, GAS_INDOOR_IND, GAS_OUTDOOR_IND) which lack TIER1 in their name."""
    return ("TIER1" in k and ("_IND_" in k or k.endswith("_IND"))) or \
           k in ("GAS_AMBIENT_IND", "GAS_INDOOR_IND", "GAS_OUTDOOR_IND")


def _is_res_key(k: str) -> bool:
    """True for Tier1 Residential threshold keys (e.g. TIER1_RES_*, TIER1_INDOOR_RES, PFAS_TIER1_RES*).
    Also covers gas thresholds (GAS_AMBIENT_RES, GAS_INDOOR_RES, GAS_OUTDOOR_RES) which lack TIER1 in their name."""
    return ("TIER1" in k and ("_RES_" in k or k.endswith("_RES"))) or \
           k in ("GAS_AMBIENT_RES", "GAS_INDOOR_RES", "GAS_OUTDOOR_RES")


def _is_vsl_key(k: str) -> bool:
    return "VSL" in k and "TIER1" not in k


# ── Style constants ───────────────────────────────────────────────────
YELLOW  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # VSL
PINK    = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Tier 1 Industrial
L_BLUE  = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Tier 1 Residential
GRAY    = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
WHITE   = PatternFill(fill_type=None)
BLUE_H  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
DARK_H  = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
THIN    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
WRAP_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_L  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
CENTER  = Alignment(horizontal="center", vertical="center")

FHE = {"name": "David",           "size": 9}   # Hebrew / numbers / dates
FEN = {"name": "Times New Roman", "size": 8}   # Pure English text


def _font(val, bold=False) -> Font:
    """
    Choose font based on cell content:
      Pure English (no Hebrew)        → Times New Roman 8
      Hebrew / mixed / numbers / dates → David 9
    """
    s = str(val) if val is not None else ""
    has_hebrew  = any('\u05D0' <= c <= '\u05EA' for c in s)
    has_english = any(c.isalpha() and c.isascii() for c in s)
    # Pure English → Times New Roman 8; everything else (Hebrew, mixed, numbers, dates) → David 9
    base = FEN if (has_english and not has_hebrew) else FHE
    return Font(**base, bold=bold)


def _round_thresh(v) -> float | None:
    """Round a threshold value to 4 significant figures for display."""
    if v is None or not isinstance(v, (int, float)):
        return v
    return _round_sf(v, 4)


def _round_sf(v, sf: int = 2):
    """Round v to sf significant figures (decimal places floored at 0).

    Examples (sf=2):
      0.809      → 0.81
      27809.52   → 27810
      0.00123    → 0.0012
    """
    if v is None or not isinstance(v, (int, float)):
        return v
    if v == 0:
        return 0
    d = math.ceil(math.log10(abs(v)))
    return round(v, max(0, sf - d))


def _num_fmt_data(val) -> str:
    """Excel number format for compound data cells.
    Whole numbers use '#,##0' to avoid a trailing decimal point.
    Fractional numbers use '#,##0.###' (up to 3 significant decimal digits).
    """
    if isinstance(val, int):
        return '#,##0'
    if isinstance(val, float) and val % 1 == 0:
        return '#,##0'
    return '#,##0.###'


def _num_fmt_thresh(val) -> str:
    """Excel number format for threshold cells.
    Whole numbers use '#,##0'; small fractions (<1) use '0.######'; others use '#,##0.####'.
    """
    if val is None:
        return 'General'
    if isinstance(val, int):
        return '#,##0'
    if isinstance(val, float) and val % 1 == 0:
        return '#,##0'
    if isinstance(val, float) and abs(val) < 1:
        return '0.######'
    return '#,##0.####'


def _fmt_lod(lod: float) -> str:
    """Format a LOD value as a clean string with no trailing zeros or decimal point.

    Examples:
      50.0  → "50"
      0.009 → "0.009"
      0.020 → "0.02"
    Uses up to 3 decimal places of precision.
    """
    if lod == int(lod):
        return str(int(lod))
    return f"{round(lod, 3):.3f}".rstrip("0").rstrip(".")


_DEPTH_PAREN_RE   = re.compile(r'^(.*?)\s*\((\d+\.?\d*)\)\s*(-?\s*DUP)?\s*$', re.IGNORECASE)
# Shimshon-2 "קק-{depth_float}-{borehole_int}[ DUP]" — depth first, borehole last
_DEPTH_FLOAT_BH_RE = re.compile(r'^(.*?)-(\d+\.\d+)-(\d+)(\s+DUP)?\s*$', re.IGNORECASE)
_DEPTH_DOT_RE     = re.compile(r'^(.*?)-(\d+\.\d+)(-DUP)?\s*$', re.IGNORECASE)
_DEPTH_DASHNUM_RE = re.compile(r'^(.*?)\s+-\s+(\d+)-(\d+)(-DUP)?\s*$', re.IGNORECASE)
# Shimshon-1 "קק-{borehole} - {depth_float}[ DUP]" — name SPACE-DASH-SPACE depth with real decimal
_DEPTH_SDS_RE     = re.compile(r'^(.*?)\s+-\s+(\d+\.?\d*)(\s+DUP)?\s*$', re.IGNORECASE)
_DEPTH_SPACE_RE   = re.compile(r'^(.*?)\s+([\d]+(?:\.[\d]+)?)\s*(?:m|מ)?$', re.IGNORECASE)
_WELL_NORM_RE     = re.compile(r'^([\u05D0-\u05EA])(\d)', re.UNICODE)

# Well letter priority for sorting: ק=0, נ=1, others=99
_WELL_LETTER_ORDER: dict[str, int] = {'ק': 0, 'נ': 1}


def _norm_borehole(s: str) -> str:
    """ק12 → ק-12, נ1 → נ-1 (add '-' after single Hebrew letter prefix).
    Also strips trailing dashes that appear in some lab report formats (e.g. 'קק-3-').
    Cross-lab transliteration: Hebrew borehole prefixes ↔ Latin equivalents
    (e.g. 'פפ-1' [Bactochem] ↔ 'PP-1' [ALS])."""
    normalized = _WELL_NORM_RE.sub(r'\1-\2', s.strip())
    normalized = normalized.rstrip('-').strip()
    return _translit_borehole_prefix(normalized)


# Hebrew ↔ Latin borehole-prefix equivalence map for cross-lab matching.
_BOREHOLE_PREFIX_EQUIV: dict[str, str] = {
    "פפ": "PP", "PP": "PP",
    "מת": "MT", "MT": "MT", "MW": "MT",
    "ק": "K",   "K": "K",
    "נ": "N",   "N": "N",
    "ד": "D",   "D": "D",
    "בה": "BH", "BH": "BH",
}


def _translit_borehole_prefix(s: str) -> str:
    """Replace a known Hebrew/Latin borehole prefix with its canonical form,
    so 'פפ-1' and 'PP-1' normalize to the same key."""
    m = re.match(r'^([A-Za-z\u05d0-\u05ea]{1,3})-(.+)$', s)
    if not m:
        return s
    prefix, rest = m.group(1), m.group(2)
    canon = _BOREHOLE_PREFIX_EQUIV.get(prefix.upper(), _BOREHOLE_PREFIX_EQUIV.get(prefix))
    if canon:
        return f"{canon}-{rest}"
    return s


def _pid_key(borehole: str) -> str:
    """Normalize borehole name for pid_map lookup.
    Strips קרקע prefix (Bactochem), maps Latin K/k to ק, removes dashes/spaces."""
    name = borehole.strip()
    if name.startswith('קרקע '):
        name = name[len('קרקע '):].strip()
    name = re.sub(r'^[Kk]-?', 'ק', name)
    return re.sub(r'[-\s]', '', name).strip()


def _pid_lookup(pid_data: dict, sid: str):
    """Depth-aware PID lookup from a raw sample ID string.

    Handles three formats:
      Format 1 (AlChem): ק-{depth}-{borehole}  e.g. 'ק-3.0-5' → bh=ק5, depth=3.0
      Format 2 (ALS):    {borehole} {depth}     e.g. 'K-2 5.0' → bh=ק2, depth=5.0
      Fallback:          last integer segment as borehole, depth=0

    Finds the shallowest pid_data[bh_key] interval where depth_to >= sample depth.
    Returns PID float (0 is valid) or '-' when nothing matches.
    """
    if not pid_data:
        return '-'
    s = sid.strip()
    # Format 1: ק-{depth}-{borehole}
    m1 = re.search(r'-(\d+(?:\.\d+)?)-(\d+)$', s)
    if m1:
        depth  = float(m1.group(1))
        bh_key = f'ק{m1.group(2)}'
    else:
        # Format 2: {borehole} {depth}  (depth is the last space-separated token)
        m2 = re.search(r'^(.+)\s+([\d.]+)$', s)
        if m2:
            bh_key = _pid_key(m2.group(1))
            depth  = float(m2.group(2))
        else:
            # Fallback: last hyphen-integer → borehole, no depth info
            m3 = re.search(r'-(\d+)$', s)
            bh_key = f'ק{m3.group(1)}' if m3 else _pid_key(s)
            depth  = 0.0

    entries = pid_data.get(bh_key, [])
    if not entries:
        return '-'
    candidates = [(d, p) for d, p in entries if d >= depth]
    if candidates:
        return min(candidates, key=lambda x: x[0])[1]
    return '-'


def _pid_lookup_split(pid_data: dict, borehole: str, depth_str: str):
    """PID lookup using pre-extracted borehole name and depth string.

    Used by portrait/landscape writers that already have split_map values,
    avoiding any ambiguity in re-parsing the composite sample-ID string.
    Covers Format 3 (TPH landscape) and all other _write_landscape rows.
    """
    if not pid_data:
        return '-'
    bh_key = _pid_key(borehole)
    try:
        depth = float(depth_str) if depth_str else 0.0
    except (ValueError, TypeError):
        depth = 0.0
    entries = pid_data.get(bh_key, [])
    if not entries:
        return '-'
    candidates = [(d, p) for d, p in entries if d >= depth]
    if candidates:
        return min(candidates, key=lambda x: x[0])[1]
    return '-'


def _borehole_sort_key(bh: str) -> tuple:
    """Sort: ק-* first, then נ-*, then others. Within each group: numeric order."""
    bh_n = _norm_borehole(bh)
    first = bh_n[0] if bh_n else ''
    priority = _WELL_LETTER_ORDER.get(first, 99)
    m = re.search(r'(\d+)', bh_n)
    num = int(m.group(1)) if m else 0
    return (priority, num, bh_n)


def _tph_sort_key(sid_depth):
    """Numeric sort key for TPH sample IDs (strings or (sid, depth) tuples).

    Extracts the first integer from the sample name for numeric ordering.
    """
    sid = sid_depth[0] if isinstance(sid_depth, tuple) else sid_depth
    nums = re.findall(r'\d+', str(sid))
    depth_part = sid_depth[1] if isinstance(sid_depth, tuple) else ""
    return (int(nums[0]) if nums else 999, depth_part)


def _dup_rich_text(bh: str):
    """
    Return CellRichText for borehole names containing 'DUP':
      Hebrew/numbers/punctuation → David 9
      'DUP' → Times New Roman 8
    Returns plain str when no DUP present.
    """
    m = re.search(r'(DUP)', bh, re.IGNORECASE)
    if not m:
        return bh
    he_if = InlineFont(rFont="David", sz=9)
    en_if = InlineFont(rFont="Times New Roman", sz=8)
    parts = []
    before = bh[:m.start()]
    after  = bh[m.end():]
    if before:
        parts.append(TextBlock(he_if, before))
    parts.append(TextBlock(en_if, m.group(1)))
    if after:
        parts.append(TextBlock(he_if, after))
    return CellRichText(*parts)


def _mixed_rich_text(s: str, bold: bool = False):
    """
    For mixed Hebrew+English strings (e.g. 'VSL קרקע'):
      English segments → Times New Roman 8
      Hebrew/other segments → David 9
    Returns CellRichText if mixed, else plain str.
    """
    has_heb = any('\u05D0' <= c <= '\u05EA' for c in s)
    has_eng = any(c.isalpha() and c.isascii() for c in s)
    if not (has_heb and has_eng):
        return s

    he_if = InlineFont(rFont="David", sz=9, b=bold)
    en_if = InlineFont(rFont="Times New Roman", sz=8, b=bold)

    segments: list[tuple[bool, str]] = []  # (is_hebrew, text)
    cur_text = ""
    cur_heb: bool | None = None

    for ch in s:
        is_heb = '\u05D0' <= ch <= '\u05EA'
        is_eng = ch.isalpha() and ch.isascii()
        ch_type = True if is_heb else (False if is_eng else None)  # None = neutral

        if ch_type is None:
            cur_text += ch
        elif ch_type != cur_heb and cur_heb is not None:
            segments.append((cur_heb, cur_text))
            cur_text = ch
            cur_heb = ch_type
        else:
            cur_text += ch
            cur_heb = ch_type

    if cur_text:
        segments.append((cur_heb if cur_heb is not None else False, cur_text))

    if len(segments) <= 1:
        return s
    return CellRichText(*[
        TextBlock(he_if if is_heb else en_if, txt) for is_heb, txt in segments
    ])


def _split_sample_depth(sid: str) -> tuple[str, str]:
    """
    Split sample ID into (borehole_name, depth_str). Handles formats:
      'ק16 (3.0)'        → ('ק-16', '3.0')
      'ק17  DUP(1.2)'    → ('ק-17', '1.2 DUP')
      'ק-16-1.2'         → ('ק-16', '1.2')
      'ק-16-1.2-DUP'     → ('ק-16', '1.2 DUP')
      'ק-16 - 1-2'       → ('ק-16', '1.2')
      'ק-16 - 1-2-DUP'   → ('ק-16', '1.2 DUP')
      'ב-1 3.0'          → ('ב-1', '3.0')
      'קק-1 - 1.5'       → ('קק-1', '1.5')   ← shimshon-1
      'קק-1.5-10'        → ('קק-10', '1.5')  ← shimshon-2
      'קק-10.0-16'       → ('קק-16', '10.0') ← shimshon-2
      'קק-3.0-14 DUP'    → ('קק-14', '3.0 DUP') ← shimshon-2 DUP
    Borehole name is always normalized (ק12 → ק-12). DUP is always attached
    to the depth (matching how SPLIT is shown), never to the borehole name.
    """
    name, depth = _split_sample_depth_raw(sid)
    m = re.match(r'^(.*?)\s+DUP\s*$', name, re.IGNORECASE)
    if m:
        stripped = m.group(1).strip()
        name = stripped if stripped else name
        depth = f"{depth} DUP".strip() if depth else "DUP"
    return name, depth


def _split_sample_depth_raw(sid: str) -> tuple[str, str]:
    s = sid.strip()
    # "name (depth)" or "name DUP(depth)"
    m = _DEPTH_PAREN_RE.match(s)
    if m:
        depth = m.group(2) + (" DUP" if m.group(3) else "")
        return _norm_borehole(m.group(1).strip()), depth

    # Shimshon-2: "prefix-depth_float-borehole_int[ DUP]"
    # e.g. "קק-1.5-10" → borehole="קק-10", depth="1.5"
    m = _DEPTH_FLOAT_BH_RE.match(s)
    if m:
        prefix   = _norm_borehole(m.group(1).strip())
        depth    = m.group(2)
        bh_num   = m.group(3)
        dup_sfx  = " DUP" if m.group(4) else ""
        borehole = f"{prefix}-{bh_num}{dup_sfx}"
        return _norm_borehole(borehole), depth

    # "name-depth[-DUP]" with real decimal point (and no trailing integer segment)
    m = _DEPTH_DOT_RE.match(s)
    if m:
        name = _norm_borehole(m.group(1).strip())
        return (name + ' DUP' if m.group(3) else name), m.group(2)

    # "name - d-d[-DUP]" dash-as-decimal (e.g. "1-2" = 1.2 m)
    m = _DEPTH_DASHNUM_RE.match(s)
    if m:
        name  = _norm_borehole(m.group(1).strip())
        depth = f"{m.group(2)}.{m.group(3)}"
        return (name + ' DUP' if m.group(4) else name), depth

    # Shimshon-1: "name - depth[ DUP]" with space-dash-space and real decimal
    # e.g. "קק-1 - 1.5" → ('קק-1', '1.5')
    # e.g. "קק-3 - 3.0 DUP" → ('קק-3 DUP', '3.0')
    m = _DEPTH_SDS_RE.match(s)
    if m:
        name = _norm_borehole(m.group(1).strip())
        if m.group(3):
            name += " DUP"
        return name, m.group(2)

    # "name depth[m]" space-separated
    m = _DEPTH_SPACE_RE.match(s)
    if m:
        return _norm_borehole(m.group(1).strip()), m.group(2).strip()
    return _norm_borehole(s), ""


# ── Threshold source footnote labels ─────────────────────────────────
_THRESHOLD_SOURCES: dict[str, str] = {
    "GAS_INDOOR_RES":        "Tier 1 RBTL Residential, Rev.7, 12/24",
    "GAS_OUTDOOR_RES":       "Tier 1 RBTL Residential, Rev.7, 12/24",
    "GAS_INDOOR_IND":        "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "GAS_OUTDOOR_IND":       "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_RES_SOIL_VH":     "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_HM_0_6": "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_HM_6":   "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_LOW":    "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_IND_SOIL_VH":     "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_HM_0_6": "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_HM_6":   "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_LOW":    "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "VSL_SOIL":              "Soil VSL, Rev. 7, 12/24",
    "GW":                    "Groundwater Standard, Rev.7, 12/24",
    "PFAS_VSL":              "PFAS VSL, Rev.7, 12/24",
    "PFAS_TIER1_RES":        "PFAS Tier 1 Residential, Rev.7, 12/24",
    "PFAS_TIER1_IND":        "PFAS Tier 1 Industrial/Commercial, Rev.7, 12/24",
    "PFAS_TIER1_RES_VERY_HIGH": "PFAS Tier 1 Residential (Very High Sensitivity), Rev.7, 12/24",
    "PFAS_TIER1_RES_0_6":      "PFAS Tier 1 Residential (High/Med, 0-6m), Rev.7, 12/24",
    "PFAS_TIER1_RES_6PLUS":    "PFAS Tier 1 Residential (High/Med, >6m), Rev.7, 12/24",
    "PFAS_TIER1_RES_NO_GW":    "PFAS Tier 1 Residential (no GW), Rev.7, 12/24",
    "PFAS_TIER1_IND_VERY_HIGH": "PFAS Tier 1 Industrial/Commercial (Very High Sensitivity), Rev.7, 12/24",
    "PFAS_TIER1_IND_0_6":      "PFAS Tier 1 Industrial/Commercial (High/Med, 0-6m), Rev.7, 12/24",
    "PFAS_TIER1_IND_6PLUS":    "PFAS Tier 1 Industrial/Commercial (High/Med, >6m), Rev.7, 12/24",
    "PFAS_TIER1_IND_NO_GW":    "PFAS Tier 1 Industrial/Commercial (no GW), Rev.7, 12/24",
}

# ── Sheet configuration ───────────────────────────────────────────────
SHEET_CONFIG: dict[str, dict] = {
    # include_lod_loq  → add LOD + LOQ columns between CAS and threshold(s) (gas sheet)
    # lod_loq_mode     → "both" adds LOD+LOQ columns, "loq" adds only LOQ (soil sheets)
    # filter_nd_safe   → exclude compounds that are ND everywhere AND LOD ≤ threshold
    # units_in_header  → embed unit in column headers; no separate יחידות column
    # include_lod_row  → add a LOD meta-row in the sample-column header band
    "SOIL_GAS_VOC": {
        "name": "גז קרקע VOC", "unit": "µg/m³",
        "include_lod_loq": True,
        "filter_nd_safe":  False,
        "units_in_header": True,
    },
    "SOIL_VOC":   {"name": "קרקע VOC",  "unit": "mg/kg", "lod_loq_mode": "both", "nd_shows_loq": True},
    "SOIL_SVOC":  {"name": "קרקע SVOC", "unit": "mg/kg", "lod_loq_mode": "both", "nd_shows_loq": True},
    "SOIL_MBTEX": {"name": "קרקע MBTEX",         "unit": "mg/kg"},
    "SOIL_TPH":   {"name": "קרקע TPH",            "unit": "mg/kg", "lod_loq_mode": "loq"},
    "SOIL_TPH_VOC":   {"name": "קרקע TPH+BTEX",      "unit": "mg/kg"},
    "SOIL_TPH_MBTEX": {"name": "קרקע TPH+MBTEX",     "unit": "mg/kg"},
    "SOIL_METALS":    {"name": "קרקע מתכות",         "unit": "mg/kg DW", "nd_shows_loq": True, "lod_loq_mode": "loq"},
    "SOIL_NUTRIENTS":  {"name": "קרקע כימיה",          "unit": "mg/kg", "nd_shows_loq": True},
    "SOIL_MICROBIOLOGY":{"name": "קרקע חיידקים",      "unit": "CFU/gr", "nd_shows_loq": True},
    "SOIL_GRAIN_SIZE": {"name": "גרנולומטריה",        "unit": "%"},
    "SOIL_PFAS":   {"name": "קרקע PFAS",       "unit": "ng/g"},
    "GW_VOC":          {"name": "מי תהום VOC",      "unit": "µg/L", "lod_loq_mode": "loq"},
    "GW_SVOC":         {"name": "מי תהום SVOC",     "unit": "µg/L", "lod_loq_mode": "loq"},
    "GW_METALS":       {"name": "מי תהום מתכות",   "unit": "µg/L", "lod_loq_mode": "loq"},
    "GW_PFAS":         {"name": "מי תהום PFAS",         "unit": "ng/L"},
    "GW_MICROBIOLOGY": {"name": "מיקרוביולוגיה מי תהום", "unit": "CFU/mL"},
    "LOWFLOW":         {"name": "pH",               "unit": ""},
    "GW_FIELD_PARAMS": {"name": "פרמטרי שדה",       "unit": ""},
}


def _ordered_unique(seq) -> list:
    seen = set()
    out = []
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


class LabReportExcel:
    """
    Build a multi-sheet Hebrew RTL Excel lab report.

    Parameters
    ----------
    records : list[dict]
        Flat list of measurement records, each with keys:
        compound, cas, sample_id, value, flag, unit, lod, analysis_type
    threshold_manager : ThresholdManager
    output_path : str
    project_name : str
    client : str
    report_date : str   (DD.MM.YYYY)
    selected_thresholds : list[str] | None
        Override which threshold keys to show.  None = use defaults per analysis.
    """

    def __init__(
        self,
        records: list[dict],
        threshold_manager: ThresholdManager,
        output_path: str = "lab_report.xlsx",
        project_name: str = "",
        client: str = "",
        report_date: str = "",
        selected_thresholds: list[str] | None = None,
        combine_tph_voc: bool = False,
        combine_tph_mbtex: bool = False,
        pid_map: dict | None = None,
        secondary_records: list[dict] | None = None,
    ):
        self.records           = records
        self.secondary_records = secondary_records or []
        self.tm                = threshold_manager
        self.out_path          = output_path
        self.project           = project_name
        self.client            = client
        self.rep_date          = report_date or date.today().strftime("%d.%m.%Y")
        self.sel_thresh        = selected_thresholds  # None → auto per analysis_type
        self.combine_tph_voc   = combine_tph_voc
        self.combine_tph_mbtex = combine_tph_mbtex
        self.pid_map           = pid_map or {}

    # ------------------------------------------------------------------
    def build(self) -> str:
        # Group records by analysis_type
        groups: dict[str, list[dict]] = defaultdict(list)
        for r in self.records:
            groups[r.get("analysis_type", "UNKNOWN")].append(r)

        # Optionally merge SOIL_TPH + SOIL_VOC into one combined sheet
        if self.combine_tph_voc and "SOIL_TPH" in groups and "SOIL_VOC" in groups:
            groups["SOIL_TPH_VOC"] = list(groups.pop("SOIL_TPH")) + list(groups.pop("SOIL_VOC"))

        # Optionally merge SOIL_TPH + SOIL_MBTEX into one combined sheet
        if self.combine_tph_mbtex and "SOIL_TPH" in groups and "SOIL_MBTEX" in groups:
            groups["SOIL_TPH_MBTEX"] = list(groups.pop("SOIL_TPH")) + list(groups.pop("SOIL_MBTEX"))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default sheet

        for atype, recs in groups.items():
            if len(recs) == 0:
                continue
            # Skip sheets with no usable compound rows (catches nan/empty after filtering)
            if not any(
                r.get("compound", "").strip() and r.get("compound", "").strip().lower() not in ("nan", "parameter", "compound", "analyte")
                for r in recs
            ):
                continue

            cfg   = SHEET_CONFIG.get(atype, {"name": atype, "unit": ""})
            sheet = wb.create_sheet(title=cfg["name"][:31])
            sheet.sheet_view.rightToLeft = True

            thresh_keys = self._thresh_keys(atype)

            if atype == "LOWFLOW":
                self._write_lowflow_sheet(sheet, recs, cfg)
            elif atype == "GW_FIELD_PARAMS":
                sec_fp = [r for r in self.secondary_records
                          if r.get("analysis_type") == "GW_FIELD_PARAMS"]
                self._write_field_params_sheet(sheet, recs, cfg, secondary_records=sec_fp)
            else:
                if not self._write_data_sheet(sheet, recs, cfg, thresh_keys):
                    wb.remove(sheet)

        # Only create directories when out_path is a real filesystem path (not BytesIO)
        if isinstance(self.out_path, (str, os.PathLike)):
            os.makedirs(os.path.dirname(self.out_path) or ".", exist_ok=True)
        # Add fuzzy threshold warnings at the bottom of each sheet
        if hasattr(self.tm, '_fuzzy_warnings') and self.tm._fuzzy_warnings:
            for ws in wb.worksheets:
                last_row = ws.max_row + 2
                for i, warning in enumerate(self.tm._fuzzy_warnings):
                    ws.cell(row=last_row + i, column=1, value=warning)
                    ws.cell(row=last_row + i, column=1).font = Font(italic=True, color="FF8C00")

        wb.save(self.out_path)
        return self.out_path

    # ------------------------------------------------------------------
    # Sheet writers
    # ------------------------------------------------------------------
    def _write_data_sheet(self, ws, records, cfg, thresh_keys):
        # When a record carries an explicit "depth" field (e.g. from readable
        # Alchem PDFs), embed it into the sample key as "name depth" so that
        # different depths of the same borehole become distinct columns.
        # _split_sample_depth already handles "name depth" space-separated format.
        # ── Secondary lab merging (two-pivot correlation) ──────────────
        # Build separate primary and secondary pivots, then correlate samples
        # by (borehole, depth) to create interleaved SPLIT columns/rows.
        sheet_atype = records[0].get("analysis_type") if records else None
        sec_records_raw = [r for r in self.secondary_records
                           if r.get("analysis_type") == sheet_atype]
        has_secondary = bool(sec_records_raw)

        def _sid(r):
            sid = r["sample_id"]
            d   = r.get("depth", "")
            return f"{sid} {d}" if d else sid

        def _match_key(sid: str) -> str:
            s = sid.strip()
            if s.startswith("קרקע "):
                s = s[len("קרקע "):].strip()
            bh, dep = _split_sample_depth(s)
            bh_norm = re.sub(r"[-\s]", "", _norm_borehole(bh))
            return f"{bh_norm}|{dep}"

        # CAS-based compound matching: built lazily when CAS numbers are available
        _cas_to_norm: dict[str, str] = {}   # cas → first-seen norm_cmp (primary wins)

        def _norm_cmp(name: str) -> str:
            """Canonical compound name for cross-lab matching.

            Handles common cross-lab name differences:
            - 'Ag - Silver' (KTE/Alchem prefix style) → strip element-symbol prefix
            - 'Aluminium' vs 'Aluminum' → normalize -ium/-um endings
            - dot-numbers vs comma-numbers: '1.1-' → '1,1-'
            """
            s = name.lower().strip()
            # dot-numbers → comma: '1.1.1-' → '1,1,1-' (use lookaround for all occurrences)
            s = re.sub(r'(?<=\d)\.(?=\d)', ',', s)
            # collapse spaces/hyphens
            s = re.sub(r'[-–\s]+', ' ', s)
            # Strip element-symbol prefix: "ag silver" → "silver", "fe iron" → "iron"
            # Pattern: one or two letters (element symbol) then a space then the rest
            m = re.match(r'^([a-z]{1,2}) ([a-z].*)$', s)
            if m:
                sym, rest = m.group(1), m.group(2)
                # Only strip if the rest is a known element name (i.e. sym is element-like)
                # Heuristic: rest is a single English word (no digit prefix)
                if re.match(r'^[a-z]+', rest) and not re.match(r'^\d', rest):
                    s = rest
            # Normalize -ium / -um endings (aluminium vs aluminum, etc.)
            s = re.sub(r'inium$', 'inum', s)
            s = re.sub(r'ium$', 'um', s)
            # chrome → chromium: normalize common synonyms
            _synonyms = {
                'chrome': 'chromum',
                'chromium': 'chromum',
                'mercury': 'mercury',
                'quicksilver': 'mercury',
            }
            # Remove parentheses content before synonym lookup
            s_clean = re.sub(r'\([^)]*\)', '', s).strip()
            s_clean = re.sub(r'\s+', ' ', s_clean)
            # TPH fraction synonyms: map ALS names to primary lab names
            _tph_synonyms = {
                'c10 c28 fraction dro': 'dro',
                'c10 c28 fraction': 'dro',
                'c10 c40 fraction tph': 'tph',
                'c24 c40 fraction oro': 'oro',
                'c24 c40 fraction': 'oro',
                'c10 c28 dro': 'dro',
                'c24 c40 oro': 'oro',
            }
            s = _tph_synonyms.get(s_clean, _tph_synonyms.get(s, s))
            return s


        # Build primary pivot
        pri_sids  = _ordered_unique(_sid(r) for r in records)
        pri_pivot: dict[str, dict] = {}   # norm_cmp → sid → entry
        pri_cmp_display: dict[str, str] = {}  # norm_cmp → original display name
        cas_map:   dict[str, str]  = {}   # keyed by display name
        lod_map:   dict[str, float | None] = {}
        loq_map:   dict[str, float | None] = {}
        sec_loq_map: dict[str, float | None] = {}
        unit_map:  dict[str, str]  = {}

        for r in records:
            cmp  = r["compound"]
            ncmp = _norm_cmp(cmp)
            sid  = _sid(r)
            if ncmp not in pri_pivot:
                pri_pivot[ncmp]       = {}
                pri_cmp_display[ncmp] = cmp
                cas_map[cmp]          = r.get("cas", "")
                unit_map[cmp]         = r.get("unit", cfg.get("unit", ""))
            pri_pivot[ncmp][sid] = (r.get("value"), r.get("flag", ""), r.get("lod"))
            if lod_map.get(cmp) is None and r.get("lod") is not None:
                lod_map[cmp] = r["lod"]
            if loq_map.get(cmp) is None and r.get("loq") is not None:
                loq_map[cmp] = r["loq"]

        # Build secondary pivot (keyed by normalized compound name)
        sec_sids   = _ordered_unique(_sid(r) for r in sec_records_raw)
        sec_pivot: dict[str, dict] = {}   # norm_cmp → sid → entry
        sec_cmp_display: dict[str, str] = {}  # norm_cmp → original secondary name

        # CAS → primary ncmp: fallback for name mismatches (e.g. "Mercury" vs "Hg - Mercury***")
        pri_cas_to_ncmp: dict[str, str] = {}
        for ncmp, display in pri_cmp_display.items():
            cas = cas_map.get(display, "")
            if cas:
                pri_cas_to_ncmp[str(cas).strip()] = ncmp

        for r in sec_records_raw:
            cmp  = r["compound"]
            ncmp = _norm_cmp(cmp)
            sid  = _sid(r)
            # CAS-based fallback: if norm name doesn't match primary but CAS does, use primary's ncmp
            sec_cas = str(r.get("cas") or "").strip()
            if ncmp not in pri_pivot and sec_cas and sec_cas in pri_cas_to_ncmp:
                ncmp = pri_cas_to_ncmp[sec_cas]
            if ncmp not in sec_pivot:
                sec_pivot[ncmp]        = {}
                sec_cmp_display[ncmp]  = cmp
                # Fill cas/unit from secondary if primary doesn't have it
                if pri_cmp_display.get(ncmp) is None:
                    pri_cmp_display[ncmp] = cmp
                    display_name = cmp
                    cas_map[display_name]  = r.get("cas", "")
                    unit_map[display_name] = r.get("unit", cfg.get("unit", ""))
            sec_pivot[ncmp][sid] = (r.get("value"), r.get("flag", ""), r.get("lod"))
            if sec_loq_map.get(ncmp) is None and r.get("loq") is not None:
                sec_loq_map[ncmp] = r["loq"]
            if lod_map.get(pri_cmp_display.get(ncmp, cmp)) is None and r.get("lod") is not None:
                lod_map[pri_cmp_display.get(ncmp, cmp)] = r["lod"]

        # Match primary sids to secondary sids by (borehole, depth) key
        sec_key_to_sid = {}
        for sid in sec_sids:
            k = _match_key(sid)
            if k not in sec_key_to_sid:
                sec_key_to_sid[k] = sid

        pri_to_sec: dict[str, str] = {}
        for sid in pri_sids:
            k = _match_key(sid)
            if k in sec_key_to_sid:
                pri_to_sec[sid] = sec_key_to_sid[k]

        SPLIT_SUFFIX = "_SPLIT"
        matched_sec_sids = set(pri_to_sec.values())

        samples = []
        for psid in pri_sids:
            samples.append(psid)
            if psid in pri_to_sec:
                samples.append(psid + SPLIT_SUFFIX)
        for ssid in sec_sids:
            if ssid not in matched_sec_sids:
                samples.append(ssid + SPLIT_SUFFIX)

        # Unified compound list (display names, union of both labs)
        all_ncmps = list(pri_pivot.keys())
        for ncmp in sec_pivot:
            if ncmp not in pri_pivot:
                all_ncmps.append(ncmp)
        compounds = [pri_cmp_display.get(ncmp, ncmp) for ncmp in all_ncmps]
        # Keep ncmp → display name mapping for pivot lookup
        _ncmp_list = all_ncmps

        # Rebuild sec_loq_map keyed by display name
        sec_loq_map = {pri_cmp_display.get(nc, nc): v for nc, v in sec_loq_map.items()}
        # Rebuild lod_map and loq_map keyed by display name (already done above for primary)

        _DASH = ("-", "dash", None)

        pivot: dict[str, dict] = {}
        for ncmp, display in zip(_ncmp_list, compounds):
            pivot[display] = {}
            for sid in samples:
                if sid.endswith(SPLIT_SUFFIX):
                    base   = sid[:-len(SPLIT_SUFFIX)]
                    sec_sid = pri_to_sec.get(base) or base
                    entry   = (sec_pivot.get(ncmp) or {}).get(sec_sid)
                    pivot[display][sid] = entry if entry is not None else _DASH
                else:
                    entry = (pri_pivot.get(ncmp) or {}).get(sid)
                    pivot[display][sid] = entry if entry is not None else _DASH

        # Enrich cas_map: for compounds with no CAS from the parser, try the
        # threshold manager's VSL tables first, then fall back to cas_lookup
        # CHEMICAL_MAP (covers metals and water-specific compounds).
        for cmp in list(cas_map):
            if not cas_map[cmp]:
                looked_up = self.tm.get_cas_by_name(cmp)
                if looked_up:
                    cas_map[cmp] = looked_up
                else:
                    resolved = _fuzzy_name_to_cas(cmp.strip().lower())
                    if resolved:
                        cas_map[cmp] = resolved

        # Get thresholds per compound (with confidence tracking)
        # uncertain_compounds: maps compound_name → threshold_table_name (for the note)
        thresh_vals: dict[str, dict[str, float | None]] = {}
        uncertain_compounds: dict[str, str] = {}  # cmp → thresh name in table
        for cmp, cas in cas_map.items():
            row_thresh = {}
            cmp_uncertain = False
            for k in thresh_keys:
                val, conf = self.tm.get_threshold_with_confidence(cas, k, compound_name=cmp)
                if conf == 'uncertain' and val is not None:
                    cmp_uncertain = True
                    row_thresh[k] = None  # treat as "no threshold" for display/coloring
                else:
                    row_thresh[k] = val
            if cmp_uncertain:
                # Look up the threshold table name for this CAS for the note
                thresh_name = getattr(self.tm, '_cas_to_thresh_name', {}).get(
                    str(cas).strip(), ""
                )
                if not thresh_name:
                    # init the cache if not yet done
                    self.tm.get_threshold_with_confidence(cas, thresh_keys[0], compound_name=cmp)
                    thresh_name = getattr(self.tm, '_cas_to_thresh_name', {}).get(
                        str(cas).strip(), ""
                    )
                uncertain_compounds[cmp] = thresh_name
            thresh_vals[cmp] = row_thresh

        # Optional: remove compounds that are ND everywhere AND LOD ≤ strictest threshold
        # (safe to exclude — cannot possibly exceed threshold)
        if cfg.get("filter_nd_safe"):
            def _should_keep(cmp: str) -> bool:
                t_limit = self._strictest(thresh_vals.get(cmp, {}))
                for sid in samples:
                    v, flag, lod = pivot.get(cmp, {}).get(sid, (None, "<LOQ", None))
                    # At least one detected value → keep
                    if flag not in ("<LOQ",) and v is not None:
                        return True
                    # ND but LOD exceeds threshold → grey → keep
                    if lod is not None and t_limit is not None and lod > t_limit:
                        return True
                return False
            compounds = [c for c in compounds if _should_keep(c)]

        # Remove category-header rows: no LOQ value AND no numeric sample values.
        # These are section labels (e.g. "Alcohols / Esters") emitted by some labs
        # that carry no analytical data and should not appear in the output sheet.
        def _is_header_row(cmp: str) -> bool:
            if loq_map.get(cmp) is not None:
                return False
            return not any(
                isinstance(v, (int, float))
                for v, _flag, _lod in pivot.get(cmp, {}).values()
            )
        compounds = [c for c in compounds if not _is_header_row(c)]
        if not compounds:
            return False

        # Remove sample columns where every compound is ND (no real data at all).
        def _sample_has_data(sid: str) -> bool:
            # For TPH, always show all samples (N.D. is a valid result)
            if cfg.get("analysis_type") == "SOIL_TPH" or all(
                r.get("analysis_type") == "SOIL_TPH" for r in records
            ):
                return True
            # For sheets where ND is shown as LOQ (e.g. VOC/BTEX), ND is a valid result
            if cfg.get("nd_shows_loq"):
                return any(
                    pivot.get(cmp, {}).get(sid) is not None
                    for cmp in compounds
                )
            return any(
                pivot.get(cmp, {}).get(sid, (None, "ND", None))[1] != "ND"
                for cmp in compounds
            )
        samples = [s for s in samples if _sample_has_data(s)]
        if not samples:
            return False

        # Per-sample metadata (soil gas: canister, sampling date, PID reading)
        sample_meta: dict[str, dict] = {}
        for r in records:
            sid = _sid(r)
            if sid not in sample_meta:
                sample_meta[sid] = {
                    "canister": r.get("canister_num", ""),
                    "date":     r.get("sampling_date", ""),
                    "pid":      r.get("pid_reading", ""),
                }

        # Auto-downgrade lod_loq_mode: when every compound has lod = None (e.g. ALS
        # reports that only provide LOR/LOQ with no separate LOD column), showing a
        # blank LOD column adds noise. Downgrade "both" → "loq" so only LOQ appears.
        cfg = copy.deepcopy(cfg)  # fresh copy each call — never mutate global SHEET_CONFIG
        if cfg.get("lod_loq_mode") == "both" and not any(
            lod_map.get(c) is not None for c in compounds
        ):
            cfg["lod_loq_mode"] = "loq"

        # depth_map: overrides depth string for a sample; populated from
        # pre-parsed record["depth_from"] when available (e.g. ALS "BH1-1.5m").
        depth_map: dict[str, str] = {}
        for r in records:
            sid = _sid(r)
            dep = r.get("depth_from")
            bh  = r.get("borehole")
            if dep is not None and sid not in depth_map:
                depth_map[sid] = str(dep)
            if bh and sid in sample_meta and not sample_meta[sid].get("borehole"):
                sample_meta[sid]["borehole"] = bh

        # Decide orientation: portrait when n_compounds >= n_samples.
        portrait = len(compounds) >= len(samples)

        header_info = {
            "project": self.project,
            "date":    self.rep_date,
            "client":  self.client,
            "unit":    cfg["unit"],
        }

        self._current_sheet_records = records
        if portrait:
            self._write_portrait(ws, compounds, samples, pivot, cas_map,
                                 lod_map, loq_map,
                                 thresh_keys, thresh_vals, header_info, cfg,
                                 sample_meta=sample_meta, unit_map=unit_map,
                                 depth_map=depth_map, pid_map=self.pid_map,
                                 uncertain_compounds=uncertain_compounds,
                                 has_secondary=has_secondary,
                                 sec_loq_map=sec_loq_map)
        else:
            self._write_landscape(ws, compounds, samples, pivot, cas_map,
                                  lod_map, loq_map,
                                  thresh_keys, thresh_vals, header_info, cfg,
                                  sample_meta=sample_meta, unit_map=unit_map,
                                  depth_map=depth_map, pid_map=self.pid_map,
                                  uncertain_compounds=uncertain_compounds,
                                  has_secondary=has_secondary,
                                  sec_loq_map=sec_loq_map)
        return True

    def _write_lowflow_sheet(self, ws, records, cfg):
        """LOWFLOW/pH: field parameters as rows, samples as columns, no thresholds.
        Extracts borehole name and depth from sample IDs (rows 2-3 metadata).
        """
        samples = _ordered_unique(r["sample_id"] for r in records)
        params  = _ordered_unique(r["compound"]  for r in records)

        pivot: dict[str, dict] = {}
        unit_map: dict[str, str] = {}
        for r in records:
            p = r["compound"]
            s = r["sample_id"]
            if p not in pivot:
                pivot[p]    = {}
                unit_map[p] = r.get("unit", "")
            v = r.get("value")
            pivot[p][s] = (round(v, 3) if isinstance(v, (int, float)) else v,
                           r.get("flag", ""), None)

        # Split sample IDs into borehole + depth
        split_map  = {sid: _split_sample_depth(sid) for sid in samples}
        boreholes  = [split_map[sid][0] for sid in samples]
        depths     = [split_map[sid][1] for sid in samples]

        N_FIXED = 2   # פרמטר | יחידות
        total_cols = N_FIXED + len(samples)

        # Row 1: merged project header (skipped when project+client both empty)
        header_written = self._write_header_row(ws, 1, total_cols)
        meta_start = 2 if header_written else 1

        # Rows meta_start…: metadata (שם קידוח, עומק [מ'])
        meta_rows = [("שם קידוח", boreholes)]
        if any(v is not None and str(v).strip() != "" for v in depths):
            meta_rows.append(("עומק [מ']", depths))
        for ri, (label, vals) in enumerate(meta_rows, meta_start):
            ws.merge_cells(start_row=ri, start_column=1,
                           end_row=ri,   end_column=N_FIXED)
            c = ws.cell(row=ri, column=1, value=label)
            c.font      = Font(**FHE, bold=True)
            c.alignment = WRAP_C
            c.border    = THIN
            ws.cell(row=ri, column=2).border = THIN
            for ci, v in enumerate(vals, N_FIXED + 1):
                cell = ws.cell(row=ri, column=ci, value=v if v else "")
                cell.border    = THIN
                cell.alignment = CENTER
                cell.font      = _font(v)

        # Column headers row (after all meta rows)
        hdr_row = meta_start + len(meta_rows)
        headers = ["פרמטר", "יחידות"] + samples
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=hdr_row, column=ci, value=h)
            c.font      = Font(**FHE, bold=True)
            c.alignment = WRAP_C
            c.border    = THIN

        # Data rows
        loq_map: dict[str, float | None] = {}
        row_num = hdr_row + 1
        for param in params:
            row_data = [param, unit_map.get(param, "")]
            for sid in samples:
                v, flag, _ = pivot.get(param, {}).get(sid, (None, "<LOQ", None))
                if flag == "ND" or flag == "nd":
                    row_data.append("N.D.")
                elif flag == "<LOQ":
                    loq_val = loq_map.get(param)
                    row_data.append(f"<{loq_val}" if loq_val is not None else "<LOQ")
                elif v is not None:
                    row_data.append(v)
                else:
                    row_data.append("-")
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font      = _font(val)
                c.alignment = CENTER
                c.border    = THIN
            row_num += 1

        # Note
        note = ws.cell(row=row_num + 1, column=1,
                       value="* ממצאי שדה בלבד, ללא השוואה לערכי סף")
        note.font = Font(**FHE, italic=True, color="808080")

        self._auto_width(ws, total_cols, hdr_row=4)

    def _write_field_params_sheet(self, ws, records, cfg, secondary_records=None):
        """פרמטרי שדה — field parameters (pH, EC, redox, etc.) per sampling location.

        Pivoted table:  פרמטר | יחידות | <sample 1> | <sample 1 SPLIT> | <sample 2> | ...
        One row per parameter, one column per distinct sample_id (plus a SPLIT
        column when the secondary lab reports field params for the same
        borehole). A sampling-date row is shown above the parameter rows
        when dates are available.
        """
        secondary_records = secondary_records or []

        # Collect distinct primary sample_ids in first-seen order
        sample_ids: list[str] = []
        seen_sids: set[str] = set()
        sample_dates: dict[str, str] = {}
        for r in records:
            sid = r.get("sample_id", "") or ""
            if sid and sid not in seen_sids:
                seen_sids.add(sid)
                sample_ids.append(sid)
            if sid and not sample_dates.get(sid):
                d = r.get("sampling_date") or r.get("date") or ""
                if d:
                    sample_dates[sid] = str(d)

        # Build secondary lookup keyed by normalized borehole name
        sec_by_borehole: dict[str, list[dict]] = defaultdict(list)
        sec_dates: dict[str, str] = {}
        for r in secondary_records:
            sid = r.get("sample_id", "") or ""
            if not sid:
                continue
            key = _norm_borehole(sid)
            sec_by_borehole[key].append(r)
            if not sec_dates.get(key):
                d = r.get("sampling_date") or r.get("date") or ""
                if d:
                    sec_dates[key] = str(d)

        has_secondary = bool(secondary_records)

        # Single-sample, no-secondary case: keep the original simple 3-column layout
        if len(sample_ids) <= 1 and not has_secondary:
            total_cols = 3
            header_written = self._write_header_row(ws, 1, total_cols)
            hdr_row = 2 if header_written else 1
            for ci, h in enumerate(["פרמטר", "יחידות", "תוצאה"], 1):
                c = ws.cell(row=hdr_row, column=ci, value=h)
                c.font      = Font(**FHE, bold=True)
                c.alignment = WRAP_C
                c.border    = THIN

            seen: set[str] = set()
            row_num = hdr_row + 1
            for r in records:
                param = r.get("compound", "").strip()
                if not param or param in seen:
                    continue
                seen.add(param)
                unit = r.get("unit", "")
                v    = r.get("value")
                display = "" if v is None else (round(v, 3) if isinstance(v, float) else v)
                row_vals = [param, unit, display]
                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=row_num, column=ci, value=val)
                    c.font      = _font(val)
                    c.border    = THIN
                    c.alignment = WRAP_L if ci == 1 else CENTER
                row_num += 1

            note = ws.cell(row=row_num + 1, column=1,
                           value="* ממצאי שדה בלבד, ללא השוואה לערכי סף")
            note.font = Font(**FHE, italic=True, color="808080")
            ws.column_dimensions["A"].width = 42
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 12
            if header_written:
                ws.row_dimensions[1].height = 20
            ws.row_dimensions[hdr_row].height = 22
            return

        # ── Multi-sample (and/or dual-lab) case: pivot with optional SPLIT cols ──
        # col_map: sample_id → (primary_col, split_col_or_None)
        col_map: dict[str, tuple[int, int | None]] = {}
        col_labels: list[tuple[int, str]] = []   # (col_index, header_label) for header row
        cur_col = 3
        for sid in sample_ids:
            key = _norm_borehole(sid)
            if key in sec_by_borehole:
                col_map[sid] = (cur_col, cur_col + 1)
                col_labels.append((cur_col, sid))
                col_labels.append((cur_col + 1, "SPLIT"))
                cur_col += 2
            else:
                col_map[sid] = (cur_col, None)
                col_labels.append((cur_col, sid))
                cur_col += 1

        # Secondary-only boreholes (no matching primary sample) get their own column
        sec_only_cols: dict[str, int] = {}
        primary_keys = {_norm_borehole(s) for s in sample_ids}
        for key in sec_by_borehole:
            if key not in primary_keys:
                sec_only_cols[key] = cur_col
                col_labels.append((cur_col, f"{key} SPLIT"))
                cur_col += 1

        total_cols = cur_col - 1
        header_written = self._write_header_row(ws, 1, total_cols)
        hdr_row = 2 if header_written else 1

        # Header row: פרמטר | יחידות | <sample columns / SPLIT>
        c = ws.cell(row=hdr_row, column=1, value="פרמטר")
        c.font = Font(**FHE, bold=True); c.alignment = WRAP_C; c.border = THIN
        c = ws.cell(row=hdr_row, column=2, value="יחידות")
        c.font = Font(**FHE, bold=True); c.alignment = WRAP_C; c.border = THIN
        for ci, label in col_labels:
            c = ws.cell(row=hdr_row, column=ci, value=label)
            c.font      = Font(**FHE, bold=True)
            c.alignment = WRAP_C
            c.border    = THIN

        date_row_written = any(sample_dates.values()) or any(sec_dates.values())
        if date_row_written:
            date_row = hdr_row + 1
            c = ws.cell(row=date_row, column=1, value="תאריך דיגום")
            c.font = Font(**FHE, italic=True); c.alignment = WRAP_C; c.border = THIN
            for sid in sample_ids:
                pcol, scol = col_map[sid]
                c = ws.cell(row=date_row, column=pcol, value=sample_dates.get(sid, ""))
                c.font = Font(**FHE, italic=True); c.alignment = CENTER; c.border = THIN
                if scol is not None:
                    key = _norm_borehole(sid)
                    c = ws.cell(row=date_row, column=scol, value=sec_dates.get(key, ""))
                    c.font = Font(**FHE, italic=True); c.alignment = CENTER; c.border = THIN
            for key, ci in sec_only_cols.items():
                c = ws.cell(row=date_row, column=ci, value=sec_dates.get(key, ""))
                c.font = Font(**FHE, italic=True); c.alignment = CENTER; c.border = THIN
            data_start_row = date_row + 1
        else:
            data_start_row = hdr_row + 1

        # Build param → unit, param → {col_index: value}
        param_order: list[str] = []
        param_unit: dict[str, str] = {}
        param_vals: dict[str, dict[int, object]] = {}

        # Field-parameter name synonyms: map lab-specific names to a canonical
        # Hebrew label so Bactochem and ALS rows for the same physical
        # parameter merge into one row instead of duplicating.
        _FP_CANON: dict[str, str] = {
            # pH
            "pH (after stabilization)":            "pH",
            "pH Value":                             "pH",
            "הגבה pH":                               "pH",
            # Electrical conductivity
            "Elctr.conductivity (after stabilization)": "מוליכות חשמלית",
            "Electrical Conductivity @ 25°C":           "מוליכות חשמלית",
            "מוליכות":                                   "מוליכות חשמלית",
            # Redox
            "Redox (after stabilization)":          "רדוקס",
            "Redox Potential":                       "רדוקס",
            "רדוקס":                                  "רדוקס",
            # Dissolved oxygen
            "Dissolved O2 (after stabilization)":   "חמצן מומס",
            "Dissolved Oxygen":                      "חמצן מומס",
            "חמצן מומס DO":                           "חמצן מומס",
            # Temperature
            "Temp (after stabilization)":           "טמפרטורה",
            "טמפרטורה":                                "טמפרטורה",
            # Turbidity
            "Turbidity (after stabilization)":      "עכירות",
            "עכירות":                                  "עכירות",
            # Sampling depth / water level
            "Sampling depth":                        "עומק דיגום",
            "עומק דיגום LOWFLOW":                      "עומק דיגום",
            "Depth of upper level":                  "מפלס עליון",
            "מפלס עליון":                              "מפלס עליון",
            "Total depth of drilling":               "עומק כללי קידוח",
            "עומק כללי קידוח":                          "עומק כללי קידוח",
        }

        def _canon_param(name: str) -> str:
            return _FP_CANON.get(name.strip(), name.strip())

        def _add_param_value(param: str, unit: str, ci: int, v):
            param = _canon_param(param)
            if param not in param_vals:
                param_vals[param] = {}
                param_order.append(param)
                param_unit[param] = unit
            display = "-" if v is None else (round(v, 3) if isinstance(v, float) else v)
            param_vals[param].setdefault(ci, display)

        for r in records:
            param = (r.get("compound") or "").strip()
            if not param:
                continue
            sid = r.get("sample_id", "") or ""
            pcol, _ = col_map.get(sid, (None, None))
            if pcol is not None:
                _add_param_value(param, r.get("unit", ""), pcol, r.get("value"))

        for sid in sample_ids:
            pcol, scol = col_map[sid]
            if scol is None:
                continue
            key = _norm_borehole(sid)
            for r in sec_by_borehole.get(key, []):
                param = (r.get("compound") or "").strip()
                if not param:
                    continue
                _add_param_value(param, r.get("unit", ""), scol, r.get("value"))

        for key, ci in sec_only_cols.items():
            for r in sec_by_borehole.get(key, []):
                param = (r.get("compound") or "").strip()
                if not param:
                    continue
                _add_param_value(param, r.get("unit", ""), ci, r.get("value"))

        row_num = data_start_row
        for param in param_order:
            c = ws.cell(row=row_num, column=1, value=param)
            c.font = _font(param); c.border = THIN; c.alignment = WRAP_L
            c = ws.cell(row=row_num, column=2, value=param_unit.get(param, ""))
            c.font = _font(param_unit.get(param, "")); c.border = THIN; c.alignment = CENTER
            for ci in range(3, total_cols + 1):
                val = param_vals[param].get(ci, "-")
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font      = _font(val)
                c.border    = THIN
                c.alignment = CENTER
            row_num += 1

        note = ws.cell(row=row_num + 1, column=1,
                       value="* ממצאי שדה בלבד, ללא השוואה לערכי סף")
        note.font = Font(**FHE, italic=True, color="808080")

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 12
        for ci in range(3, total_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14
        if header_written:
            ws.row_dimensions[1].height = 20
        ws.row_dimensions[hdr_row].height = 22

    # ------------------------------------------------------------------
    # Portrait layout: compounds as rows
    #
    # Standard mode (soil / GW):
    #   A: compound | B: CAS | C…: threshold col(s) | next: יחידות | then: samples
    #
    # LOD/LOQ mode (soil-gas, cfg["include_lod_loq"]=True):
    #   A: compound | B: CAS | C: LOD [unit] | D: LOQ [unit]
    #   | E…: threshold col(s) | then: samples  (no separate יחידות col)
    # ------------------------------------------------------------------
    def _write_portrait(self, ws, compounds, samples, pivot, cas_map,
                        lod_map, loq_map,
                        thresh_keys, thresh_vals, hinfo, cfg=None, sample_meta=None,
                        unit_map=None, depth_map=None, pid_map=None,
                        uncertain_compounds=None,
                        has_secondary=False, sec_loq_map=None):
        cfg         = cfg or {}
        sample_meta = sample_meta or {}
        uncertain_compounds = uncertain_compounds or {}
        sec_loq_map = sec_loq_map or {}
        # Separate primary vs secondary sample IDs
        pri_samples = [s for s in samples if "_SEC" not in s]
        sec_samples = [s for s in samples if "_SEC" in s]
        unit            = hinfo["unit"]
        include_lod_loq = cfg.get("include_lod_loq", False)   # gas sheet: full LOD+LOQ mode
        lod_loq_mode    = cfg.get("lod_loq_mode", False)       # soil: "both" or "loq"
        units_in_header = cfg.get("units_in_header", False)

        N_COMPOUND = 2                         # A: compound, B: CAS Number
        if include_lod_loq or lod_loq_mode == "both":
            N_LOD_LOQ = 2 + (1 if has_secondary else 0)  # LOD + LOQ ראשית [+ LOQ משנית]
        elif lod_loq_mode == "loq":
            N_LOD_LOQ = 1 + (1 if has_secondary else 0)  # LOQ ראשית [+ LOQ משנית]
        else:
            N_LOD_LOQ = 0
        N_THRESH   = len(thresh_keys)
        N_UNIT     = 0                             # unit shown in "Final conc." header
        N_FIXED    = N_COMPOUND + N_LOD_LOQ + N_THRESH + N_UNIT
        total_cols = N_FIXED + len(samples)

        thresh_labels = [THRESHOLD_LABELS.get(k, k) for k in thresh_keys]

        if include_lod_loq:
            # ── Rows 1-N: sample metadata rows (soil gas) ─────────────
            meta_rows = [
                ("שם קידוח",           [s for s in samples]),
                ("תאריך ביצוע הדיגום", [sample_meta.get(s, {}).get("date",     "") for s in samples]),
                ("מספר קניסטר",        [sample_meta.get(s, {}).get("canister", "") for s in samples]),
            ]
            if pid_map:
                _gas_pids = [_pid_lookup(pid_map, s) for s in samples]
                meta_rows.append(('קריאת PID [ppm]', _gas_pids))
            for ri, (label, vals) in enumerate(meta_rows, 1):
                # Merge label across all fixed columns (A → last fixed col)
                ws.merge_cells(start_row=ri, start_column=1,
                               end_row=ri,   end_column=N_FIXED)
                if "PID" in label:
                    # Rich text: Hebrew parts → David 9 bold; "PID" → Times New Roman 8 bold
                    he_if = InlineFont(rFont="David", sz=9, b=True)
                    en_if = InlineFont(rFont="Times New Roman", sz=8, b=True)
                    parts = label.split("PID")
                    c = ws.cell(row=ri, column=1)
                    c.value = CellRichText(
                        TextBlock(he_if, parts[0]),
                        TextBlock(en_if, "PID"),
                        TextBlock(he_if, parts[1]),
                    )
                    c.font = Font(**FHE, bold=True)
                else:
                    c = ws.cell(row=ri, column=1, value=label)
                    c.font = _font(label, bold=True)
                c.alignment = WRAP_C
                c.border    = THIN
                # No fill on label or fixed columns
                for ci in range(2, N_FIXED + 1):
                    ws.cell(row=ri, column=ci).border = THIN
                # Sample value cells
                is_date_row = "תאריך" in label
                for ci, v in enumerate(vals, N_FIXED + 1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.border    = THIN
                    cell.alignment = CENTER
                    if is_date_row:
                        # Leave empty for manual entry; apply date format
                        cell.number_format = "DD/MM/YYYY"
                    else:
                        cell.value = v
                        cell.font  = _font(v)
            # ── Column headers row (no fill; sample cols merged) ────────
            lod_hdr = f"LOD [{unit}]"
            loq_hdr = f"LOQ [{unit}]"
            headers = (["תרכובת", "CAS Number", lod_hdr, loq_hdr]
                       + thresh_labels
                       + [""] * len(samples))
            hdr_row = len(meta_rows) + 1
        else:
            # ── Row 1: merged project info header (skipped when project+client empty) ──
            header_written = self._write_header_row(ws, 1, total_cols, hinfo)
            meta_start = 2 if header_written else 1
            # ── Sort samples and build metadata ────────────────────────
            split_p = {sid: _split_sample_depth(sid) for sid in samples}
            if depth_map:
                split_p = {sid: (split_p[sid][0], depth_map.get(sid, split_p[sid][1]))
                           for sid in samples}
            # Override borehole from pre-parsed record fields when _split_sample_depth
            # could not parse the ID format (e.g. ALS "BH1-1.5m" style).
            for sid in samples:
                bh_override = sample_meta.get(sid, {}).get("borehole")
                if bh_override:
                    split_p[sid] = (_norm_borehole(bh_override), split_p[sid][1])
            # Sort samples: SPLIT follows its paired primary
            def _sec_sort_key(sid):
                is_split = sid.endswith("_SPLIT")
                base = sid[:-6] if is_split else sid
                bh, dep = split_p.get(base, split_p.get(sid, ("", "")))
                return (*_borehole_sort_key(bh), float(dep) if dep else 0.0, 1 if is_split else 0)
            samples = sorted(samples, key=_sec_sort_key)

            # Build display values: SPLIT samples show same borehole, depth="3.0 SPLIT"
            split_sec = {}
            for sid in samples:
                if sid.endswith("_SPLIT"):
                    base = sid[:-6]
                    bh, dep = split_p.get(base, _split_sample_depth(base))
                    split_sec[sid] = (bh, f"{dep} SPLIT" if dep else "SPLIT")
                else:
                    split_sec[sid] = split_p.get(sid, _split_sample_depth(sid))

            boreholes = [_dup_rich_text(split_sec[sid][0]) for sid in samples]
            depths    = [split_sec[sid][1] for sid in samples]
            meta_rows = [("שם קידוח", boreholes)]
            if any(v is not None and str(v).strip() != "" for v in depths):
                meta_rows.append(("עומק [מ']", depths))
            if pid_map:
                pid_vals_pm = [_pid_lookup_split(pid_map, split_p[sid][0], split_p[sid][1])
                               for sid in samples]
                meta_rows.append(("קריאת PID [ppm]", pid_vals_pm))
            if cfg.get("include_lod_row"):
                def _min_sample_lod(sid):
                    vals = [pivot[c][sid][2] for c in compounds if sid in pivot.get(c, {})]
                    vals = [v for v in vals if v is not None]
                    return _round_sf(min(vals)) if vals else ""
                meta_rows.append((f"LOD [{unit}]", [_min_sample_lod(sid) for sid in samples]))
            for ri, (label, vals) in enumerate(meta_rows, meta_start):
                ws.merge_cells(start_row=ri, start_column=1,
                               end_row=ri,   end_column=N_FIXED)
                c = ws.cell(row=ri, column=1, value=label)
                c.font      = _font(label, bold=True)
                c.alignment = WRAP_C
                c.border    = THIN
                for ci in range(2, N_FIXED + 1):
                    ws.cell(row=ri, column=ci).border = THIN
                for ci, v in enumerate(vals, N_FIXED + 1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.border    = THIN
                    cell.alignment = CENTER
                    if v != "" and v is not None:
                        cell.value = v
                        if not isinstance(v, CellRichText):
                            cell.font = _font(v)
            # ── Column headers row (after all meta rows) ───────────────
            hdr_row = meta_start + len(meta_rows)
            if lod_loq_mode == "both":
                loq_pri_lbl = f"LOQ ראשית [{unit}]" if has_secondary else f"LOQ [{unit}]"
                lod_loq_hdrs = [f"LOD [{unit}]", loq_pri_lbl]
                if has_secondary:
                    lod_loq_hdrs.append(f"LOQ משנית [{unit}]")
            elif lod_loq_mode == "loq":
                loq_pri_lbl = f"LOQ ראשית [{unit}]" if has_secondary else f"LOQ [{unit}]"
                lod_loq_hdrs = [loq_pri_lbl]
                if has_secondary:
                    lod_loq_hdrs.append(f"LOQ משנית [{unit}]")
            else:
                lod_loq_hdrs = []
            headers = (["תרכובת", "CAS Number"]
                       + lod_loq_hdrs
                       + thresh_labels
                       + samples)

        # Write fixed column headers (no fill on any)
        for ci, h in enumerate(headers[:N_FIXED], 1):
            rv = _mixed_rich_text(h, bold=True) if isinstance(h, str) else h
            c = ws.cell(row=hdr_row, column=ci, value=rv)
            c.font      = _font(h, bold=True)
            c.alignment = WRAP_C
            c.border    = THIN

        # Sample column headers
        if len(samples) > 0:
            sample_start = N_FIXED + 1
            sample_end   = N_FIXED + len(samples)
            if include_lod_loq:
                # Gas format: sample IDs already in meta rows → merge header with "Final conc."
                if sample_end > sample_start:
                    ws.merge_cells(start_row=hdr_row, start_column=sample_start,
                                   end_row=hdr_row,   end_column=sample_end)
                conc_hdr = ws.cell(row=hdr_row, column=sample_start,
                                   value=f"Final conc. [{unit}]")
                conc_hdr.font      = _font(f"Final conc. [{unit}]", bold=True)
                conc_hdr.alignment = CENTER
                conc_hdr.border    = THIN
                for ci in range(sample_start + 1, sample_end + 1):
                    ws.cell(row=hdr_row, column=ci).border = THIN
            else:
                # Soil/GW format: individual sample IDs as column headers.
                # Use borehole name only (split_p[sid][0]) when depth is embedded
                # in the composite key — depth is already shown in the עומק row.
                for ci, sid in enumerate(samples, sample_start):
                    display = split_sec[sid][0] if split_sec and sid in split_sec else (
                        split_p[sid][0] if split_p and sid in split_p else sid
                    )
                    c = ws.cell(row=hdr_row, column=ci, value=display)
                    c.font      = _font(display, bold=True)
                    c.alignment = CENTER
                    c.border    = THIN

        # ── Data rows ─────────────────────────────────────────────────
        data_row = hdr_row + 1
        has_gray = False   # tracks whether any gray-filled cell was written

        for cmp in compounds:
            cas    = cas_map.get(cmp, "")
            t_vals = thresh_vals.get(cmp, {})

            lod_val = lod_map.get(cmp)
            loq_val = loq_map.get(cmp)
            _empty = "-" if has_secondary else ""
            lod_disp = _round_sf(lod_val) if isinstance(lod_val, float) else (lod_val if lod_val is not None else _empty)
            loq_disp = _round_sf(loq_val) if isinstance(loq_val, float) else (loq_val if loq_val is not None else _empty)

            # Threshold values — uncertain ones already set to None in thresh_vals
            thresh_row = [
                _round_thresh(t_vals.get(k)) if _round_thresh(t_vals.get(k)) is not None
                else "לא קיים"
                for k in thresh_keys
            ]

            # Sample values — build display strings + keep raw for colouring
            sample_vals: list = []
            for sid in samples:
                entry = pivot.get(cmp, {}).get(sid, (None, "<LOQ", None))
                v, flag, lod = entry
                # Dash sentinel: compound not analysed by this lab for this sample
                if flag == "dash":
                    sample_vals.append(("-", None, "dash", None))
                    continue
                if flag == "<LOQ":
                    loq_ref = loq_val or v
                    display = _round_sf(loq_ref) if isinstance(loq_ref, float) else None
                elif flag == "<LOD":
                    display = f"<{_fmt_lod(lod)}" if lod is not None else None
                elif flag == "<":
                    display = f"<{v}" if isinstance(v, float) else f"<{v}"
                elif flag == "ND":
                    display = "N.D."
                else:
                    display = v
                sample_vals.append((display, v, flag, lod))

            if include_lod_loq or lod_loq_mode == "both":
                sec_loq_disp = _round_sf(sec_loq_map.get(cmp)) if has_secondary and sec_loq_map.get(cmp) is not None else ("-" if has_secondary else None)
                fixed_vals = [cmp, cas, lod_disp, loq_disp] + ([sec_loq_disp] if has_secondary else [])
            elif lod_loq_mode == "loq":
                sec_loq_disp = _round_sf(sec_loq_map.get(cmp)) if has_secondary and sec_loq_map.get(cmp) is not None else ("-" if has_secondary else None)
                fixed_vals = [cmp, cas, loq_disp] + ([sec_loq_disp] if has_secondary else [])
            else:
                fixed_vals = [cmp, cas]
            row_data = fixed_vals + thresh_row + [sv[0] for sv in sample_vals]

            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=data_row, column=ci, value=val)
                c.font      = _font(val)
                c.alignment = WRAP_C if ci == 1 else CENTER
                c.border    = THIN

                # ── LOD / LOQ columns ──────────────────────────────────
                if N_LOD_LOQ and N_COMPOUND < ci <= N_COMPOUND + N_LOD_LOQ:
                    pass   # no special fill, just left as-is

                # ── Threshold columns: no fill ─────────────────────────
                elif (N_COMPOUND + N_LOD_LOQ) < ci <= (N_COMPOUND + N_LOD_LOQ + N_THRESH):
                    if val == "לא קיים":
                        c.font = Font(**FHE, color="000000", italic=False)
                    else:
                        c.font          = Font(**FHE)   # David 9, no bold
                        c.number_format = _num_fmt_thresh(val)
                        c.alignment     = CENTER

                # ── Sample columns: colour coding ──────────────────────
                elif ci > N_FIXED:
                    si = ci - N_FIXED - 1
                    display, num_v, flag, lod = sample_vals[si]
                    if flag == "<LOQ" and isinstance(val, (int, float)):
                        c.number_format = '"<"0.0##'
                    # If threshold is uncertain, skip exceedance colouring entirely
                    if cmp not in uncertain_compounds and flag != "dash":
                        vsl_lim      = self._vsl_limit(t_vals)
                        tier1_ind    = self._tier1_ind_limit(t_vals)
                        tier1_res    = self._tier1_res_limit(t_vals)
                        any_lim      = self._strictest(t_vals)
                        if any_lim is not None:
                            _nd_at_loq = flag == "ND"
                            # GREY first: ND is a non-detection regardless of stored value
                            if _nd_at_loq:
                                if loq_val is not None and loq_val > any_lim:
                                    c.font = _font(display, bold=True)
                                    c.number_format = '"<"0.0##'
                                    has_gray = True
                            # BOLD: threshold < LOD/LOQ → uncertain exclusion
                            elif flag in ("<LOD", "<LOQ", "<"):
                                lod_num = (
                                    lod     if lod     is not None else
                                    lod_val if lod_val is not None else
                                    loq_val if loq_val is not None else
                                    (num_v  if isinstance(num_v, (int, float)) else None)
                                )
                                if lod_num is not None and lod_num > any_lim:
                                    c.font = _font(display, bold=True)
                                    has_gray = True
                            # COLOUR: real detection vs threshold
                            elif isinstance(num_v, (int, float)):
                                if tier1_ind is not None and num_v > tier1_ind:
                                    c.fill = PINK      # exceeds Tier 1 Industrial
                                    c.font = Font(**FHE, bold=True)
                                elif tier1_res is not None and num_v > tier1_res:
                                    c.fill = L_BLUE    # exceeds Tier 1 Residential
                                    c.font = Font(**FHE, bold=True)
                                elif vsl_lim is not None and num_v > vsl_lim:
                                    c.fill = YELLOW    # exceeds VSL only
                                    c.font = Font(**FHE, bold=True)

            # ── Uncertain threshold: asterisk in compound name + note ──
            if cmp in uncertain_compounds:
                # Append asterisk to the compound name cell
                name_cell = ws.cell(row=data_row, column=1)
                if name_cell.value and not str(name_cell.value).endswith(' *'):
                    name_cell.value = str(name_cell.value) + ' *'
                # Write the note in the first column after all data
                note_col = N_FIXED + len(samples) + 1
                thresh_name = uncertain_compounds[cmp]
                note_text = (
                    f"קיים ערך סף לתרכובת דומה בשם: {thresh_name}"
                    if thresh_name else
                    "קיים ערך סף לתרכובת דומה — נדרש לבדוק בטבלת ערכי הסף"
                )
                note_cell = ws.cell(row=data_row, column=note_col, value=note_text)
                note_cell.font = Font(**FHE, italic=True)
                note_cell.alignment = CENTER
                note_cell.border = THIN

            data_row += 1

        # ── TOTAL TPH row (DRO + ORO, only for TPH sheets) ────────────
        is_tph_sheet = (
            cfg.get("analysis_type") == "SOIL_TPH"
            or all(r.get("analysis_type") == "SOIL_TPH" for r in self._current_sheet_records)
        )
        DRO_NAMES = {"c10 - c28 fraction (dro)", "dro", "c10-c28 (dro)"}
        ORO_NAMES = {"c24 - c40 fraction (oro)", "oro", "c24-c40 (oro)"}
        has_dro = any(c.lower() in DRO_NAMES for c in compounds)
        has_oro = any(c.lower() in ORO_NAMES for c in compounds)
        # Only add if input doesn't already contain a "total" TPH compound
        has_total_already = any("total" in c.lower() and "tph" in c.lower() or
                                "total petroleum" in c.lower() for c in compounds)

        if is_tph_sheet and (has_dro or has_oro) and not has_total_already:
            dro_key = next((c for c in compounds if c.lower() in DRO_NAMES), None)
            oro_key = next((c for c in compounds if c.lower() in ORO_NAMES), None)
            dro_loq = loq_map.get(dro_key) or 0
            oro_loq = loq_map.get(oro_key) or 0
            total_loq = dro_loq + oro_loq

            n_fixed_cols = len(fixed_vals) if compounds else (N_COMPOUND + N_LOD_LOQ + N_THRESH)

            # VSL threshold for Total TPH (350 mg/kg standard)
            TPH_THRESH = 350.0

            total_sample_vals = []
            for sid in samples:
                dro_v, dro_f, _ = pivot.get(dro_key, {}).get(sid, (None, "<LOQ", None)) if dro_key else (None, "<LOQ", None)
                oro_v, oro_f, _ = pivot.get(oro_key, {}).get(sid, (None, "<LOQ", None)) if oro_key else (None, "<LOQ", None)

                dro_num = dro_v if isinstance(dro_v, (int, float)) else (dro_loq if dro_key else 0)
                oro_num = oro_v if isinstance(oro_v, (int, float)) else (oro_loq if oro_key else 0)
                both_below = dro_f in ("<LOQ", "<LOD", "<") and oro_f in ("<LOQ", "<LOD", "<")

                total_num = dro_num + oro_num
                if both_below:
                    display = total_num  # will be formatted as "<X"
                    flag = "<LOQ"
                else:
                    display = total_num
                    flag = ""
                total_sample_vals.append((display, total_num, flag))

            # Build row
            if lod_loq_mode == "loq":
                row_data = ["Total TPH", "DRO+ORO", _round_sf(total_loq)]
                if has_secondary:
                    row_data.append("")   # secondary LOQ placeholder for Total TPH
            elif include_lod_loq or lod_loq_mode == "both":
                row_data = ["Total TPH", "DRO+ORO", "", _round_sf(total_loq)]
                if has_secondary:
                    row_data.append("")   # secondary LOQ placeholder
            else:
                row_data = ["Total TPH", "DRO+ORO"]

            # Threshold placeholders
            row_data += ["לא קיים"] * N_THRESH
            row_data += [sv[0] for sv in total_sample_vals]

            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=data_row, column=ci, value=val)
                c.font      = Font(**FHE, bold=True)
                c.alignment = WRAP_C if ci == 1 else CENTER
                c.border    = THIN
                if ci > N_FIXED:
                    si = ci - N_FIXED - 1
                    display, total_num, flag = total_sample_vals[si]
                    if flag == "<LOQ" and isinstance(val, (int, float)):
                        c.number_format = '"<"0.0##'
                    elif isinstance(total_num, (int, float)) and total_num > TPH_THRESH:
                        c.fill = YELLOW
            data_row += 1
        n_legend = self._write_legend(ws, data_row + 1, include_gray=has_gray,
                                      thresh_keys=thresh_keys)
        # ── Threshold source footnotes (only for keys with ≥1 defined value) ──
        active_keys = [k for k in thresh_keys
                       if any(thresh_vals.get(c, {}).get(k) is not None for c in compounds)]
        note_row = data_row + 1 + n_legend + 1
        for note in self._threshold_source_notes(active_keys):
            c = ws.cell(row=note_row, column=1, value=f"* {note}")
            c.font = Font(**FEN, italic=True, color="808080")
            c.fill = WHITE
            note_row += 1
        if include_lod_loq:
            c = ws.cell(row=note_row, column=1,
                        value="* ספי חש מוגדרים לפי תקנות איכות אויר")
            c.font = Font(**FHE, italic=True, color="808080")
            c.fill = WHITE
            note_row += 1
        # ── Secondary lab footnote ──────────────────────────────────────
        if has_secondary:
            c = ws.cell(row=note_row, column=1,
                        value="SPLIT = תוצאות מעבדה משנית")
            c.font = Font(**FHE, italic=True, color="808080")
            c.fill = WHITE
        self._auto_width(ws, N_FIXED + len(samples), hdr_row=hdr_row)

    # ------------------------------------------------------------------
    # Landscape layout: samples as rows (when n_samples > n_compounds)
    # ------------------------------------------------------------------
    def _write_landscape(self, ws, compounds, samples, pivot, cas_map,
                         lod_map, loq_map,
                         thresh_keys, thresh_vals, hinfo, cfg=None, sample_meta=None,
                         unit_map=None, depth_map=None, pid_map=None,
                         uncertain_compounds=None,
                         has_secondary=False, sec_loq_map=None):
        cfg      = cfg or {}
        unit_map = unit_map or {}
        uncertain_compounds = uncertain_compounds or {}
        sec_loq_map = sec_loq_map or {}

        # Separate primary vs secondary samples for display
        pri_samples  = [s for s in samples if not s.endswith("_SPLIT")]
        sec_samples  = [s for s in samples if s.endswith("_SPLIT")]

        # ── Depth detection & sample sorting ────────────────────────────
        split_map = {}
        for sid in samples:
            if sid.endswith("_SPLIT"):
                base = sid[:-6]
                bh, dep = _split_sample_depth(base)
                split_map[sid] = (bh, "SPLIT")
            else:
                split_map[sid] = _split_sample_depth(sid)
        if depth_map:
            split_map = {sid: (split_map[sid][0], depth_map.get(sid, split_map[sid][1]))
                         for sid in samples}
        # Override borehole from pre-parsed record fields when available.
        if sample_meta:
            for sid in samples:
                bh_override = sample_meta.get(sid, {}).get("borehole")
                if bh_override:
                    split_map[sid] = (_norm_borehole(bh_override), split_map[sid][1])
        has_depth = any(depth for _, depth in split_map.values())

        if has_depth:
            # Sort samples: ק first, נ second, others last; within group by number then depth
            def _sort_key(sid):
                bh, dep = split_map[sid]
                dep_clean = dep.replace('SPLIT', '').strip() if dep else ''
                is_split = 1 if sid.endswith("_SPLIT") else 0
                return (*_borehole_sort_key(bh), float(dep_clean) if dep_clean else 0.0, is_split)
            samples = sorted(samples, key=_sort_key)

        # Column count: borehole + depth (when present) + PID (from pid_map) + compounds
        has_pid       = bool(pid_map)
        depth_offset  = 1 if has_depth else 0
        pid_offset    = 1 if has_pid   else 0
        total_cols    = 1 + depth_offset + pid_offset + len(compounds)
        cmp_col_start = 2 + depth_offset + pid_offset  # 1-based col of first compound

        # ── Row 1: merged project header (skipped when project+client both empty) ──
        header_written = self._write_header_row(ws, 1, total_cols, hinfo)
        hdr_base = 2 if header_written else 1

        # ── Rows hdr_base to hdr_base+2: compound names / CAS / unit ───
        _pid_hdr = ["PID [ppm]"] if has_pid else []
        _pid_pad = [""]         if has_pid else []
        if has_depth:
            row2_data = ["שם קידוח", "עומק [מ']"] + _pid_hdr + [
                (c + ' *' if c in uncertain_compounds else c) for c in compounds
            ]
            row3_data = ["CAS Number", ""]          + _pid_pad + [cas_map.get(c, "") for c in compounds]
            row4_data = ["יחידות", ""]              + _pid_pad + [unit_map.get(c, hinfo["unit"]) for c in compounds]
        else:
            row2_data = ["שם קידוח"] + _pid_hdr + [
                (c + ' *' if c in uncertain_compounds else c) for c in compounds
            ]
            row3_data = ["CAS Number"] + _pid_pad + [cas_map.get(c, "") for c in compounds]
            row4_data = ["יחידות"]     + _pid_pad + [unit_map.get(c, hinfo["unit"]) for c in compounds]

        for ri, row_vals in enumerate([row2_data, row3_data, row4_data], hdr_base):
            for ci, v in enumerate(row_vals, 1):
                rv = _mixed_rich_text(v, bold=True) if isinstance(v, str) else v
                c = ws.cell(row=ri, column=ci, value=rv)
                c.font      = _font(v, bold=True)
                c.alignment = WRAP_C
                c.border    = THIN
                # No fill on header rows 2-4 (rows 1-4 are fill-free)

        # ── Optional LOQ header row (per-compound, before thresholds) ──
        lod_loq_mode = (cfg or {}).get("lod_loq_mode", False)
        data_row = hdr_base + 3
        if lod_loq_mode:
            unit     = hinfo["unit"]
            loq_lbl  = f"LOQ ראשית [{unit}]" if has_secondary else f"LOQ [{unit}]"
            lc = ws.cell(row=data_row, column=1, value=loq_lbl)
            lc.font      = _font(loq_lbl, bold=True)
            lc.alignment = WRAP_C
            lc.border    = THIN
            for fc in range(2, cmp_col_start):
                ws.cell(row=data_row, column=fc).border = THIN
            for ci, cmp in enumerate(compounds, cmp_col_start):
                loq_val  = loq_map.get(cmp)
                loq_disp = _round_sf(loq_val) if isinstance(loq_val, float) else ""
                c = ws.cell(row=data_row, column=ci)
                c.value     = loq_disp
                c.font      = _font(loq_disp)
                c.alignment = CENTER
                c.border    = THIN
            data_row += 1

            # ── Secondary LOQ row (when secondary lab is present) ──
            if has_secondary:
                sec_loq_lbl = f"LOQ משנית [{unit}]"
                sc = ws.cell(row=data_row, column=1, value=sec_loq_lbl)
                sc.font      = _font(sec_loq_lbl, bold=True)
                sc.alignment = WRAP_C
                sc.border    = THIN
                for fc in range(2, cmp_col_start):
                    ws.cell(row=data_row, column=fc).border = THIN
                for ci, cmp in enumerate(compounds, cmp_col_start):
                    sloq_val  = sec_loq_map.get(cmp)
                    sloq_disp = _round_sf(sloq_val) if isinstance(sloq_val, float) else "-"
                    c = ws.cell(row=data_row, column=ci)
                    c.value     = sloq_disp
                    c.font      = _font(sloq_disp)
                    c.alignment = CENTER
                    c.border    = THIN
                data_row += 1

        # ── Threshold rows (BEFORE sample data) ─────────────────────────
        UNDEF_FONT  = Font(**FHE, color="000000", italic=False)
        for tk in thresh_keys:
            label    = THRESHOLD_LABELS.get(tk, tk)
            # Use plain string with readingOrder=2 so Excel treats the paragraph as RTL
            # even when the label starts with an English acronym like "TIER1".
            lbl_cell = ws.cell(row=data_row, column=1, value=label)
            lbl_cell.font      = _font(label, bold=False)
            lbl_cell.border    = THIN
            lbl_cell.alignment = Alignment(horizontal="right", vertical="center",
                                           wrap_text=True, readingOrder=2)
            # Fill fixed cols (depth + PID) in threshold rows
            for fc in range(2, cmp_col_start):
                ws.cell(row=data_row, column=fc).border = THIN
            for ci, cmp in enumerate(compounds, cmp_col_start):
                cas  = cas_map.get(cmp, "")
                tval = _round_thresh(thresh_vals.get(cmp, {}).get(tk))
                c = ws.cell(row=data_row, column=ci)
                c.border = THIN
                is_uncertain_cmp = cmp in uncertain_compounds
                if tval is None:
                    c.value     = "לא קיים"
                    c.font      = UNDEF_FONT
                    c.alignment = CENTER
                elif is_uncertain_cmp:
                    c.value     = "לא קיים"
                    c.font      = UNDEF_FONT
                    c.alignment = CENTER
                else:
                    c.value         = tval
                    c.font          = Font(**FHE)
                    c.number_format = _num_fmt_thresh(tval)
                    c.alignment     = CENTER
            data_row += 1

        # ── Sample data rows ─────────────────────────────────────────────
        first_sample_row = data_row   # remember for borehole-merge pass
        has_gray = False              # tracks whether any gray-filled cell was written

        for sid in samples:
            borehole, depth_str = split_map[sid]
            row_meta: list[tuple] = []
            col_vals: list = []

            # SPLIT rows: borehole=same as primary, depth="3.0 SPLIT" (matching image format)
            if has_secondary and depth_str == "SPLIT":
                base_sid = sid[:-6] if sid.endswith("_SPLIT") else sid
                base_bh, base_dep = split_map.get(base_sid, _split_sample_depth(base_sid))
                bh_cell_val = _dup_rich_text(base_bh)
                depth_display = f"{base_dep} SPLIT" if base_dep else "SPLIT"
            else:
                bh_cell_val = _dup_rich_text(borehole)
                depth_display = depth_str

            pid_cell = _pid_lookup_split(pid_map, borehole, depth_str) if has_pid else None
            if has_depth:
                col_vals = [bh_cell_val, depth_display if depth_display else ""] + ([pid_cell] if has_pid else [])
            else:
                col_vals = [bh_cell_val] + ([pid_cell] if has_pid else [])

            for cmp in compounds:
                entry = pivot.get(cmp, {}).get(sid, (None, "<LOQ", None))
                v, flag, lod = entry
                loq_val = loq_map.get(cmp)
                is_split_row = sid.endswith("_SPLIT")
                if flag == "dash":
                    display = "-"
                elif flag == "<LOQ":
                    loq_ref = loq_val or v
                    if is_split_row:
                        # Use secondary LOQ for display
                        sec_loq = sec_loq_map.get(cmp) if sec_loq_map else None
                        ref = sec_loq if sec_loq is not None else loq_ref
                        display = f"<{_round_sf(ref)}" if isinstance(ref, (int, float)) else f"<{ref}"
                    else:
                        display = _round_sf(loq_ref) if isinstance(loq_ref, float) else None
                elif flag == "<LOD":
                    display = f"<{_fmt_lod(lod)}" if lod is not None else None
                elif flag == "<":
                    loq_ref = loq_val or v
                    display = f"<{loq_ref}" if isinstance(loq_ref, float) else f"<{loq_ref}"
                elif flag == "ND":
                    display = "N.D."
                else:
                    display = v
                col_vals.append(display)
                row_meta.append((v, flag, lod))

            for ci, val in enumerate(col_vals, 1):
                c = ws.cell(row=data_row, column=ci, value=val)
                # CellRichText carries its own fonts; plain values use _font()
                if not isinstance(val, CellRichText):
                    c.font = _font(val)
                # Number format for numeric compound values
                if ci >= cmp_col_start and isinstance(val, (int, float)):
                    comp_idx               = ci - cmp_col_start
                    _, flag_nf, _          = row_meta[comp_idx]
                    if flag_nf == "<LOQ":
                        c.number_format = '"<"0.0##'
                    else:
                        c.number_format = _num_fmt_data(val)
                c.alignment = CENTER
                c.border    = THIN

                if ci >= cmp_col_start:
                    comp_idx               = ci - cmp_col_start
                    cmp_name               = compounds[comp_idx]
                    num_v, flag_cell, lod_cell = row_meta[comp_idx]
                    t_vals                 = thresh_vals.get(cmp_name, {})
                    # Skip exceedance colouring for uncertain compounds or dash cells
                    if cmp_name in uncertain_compounds or flag_cell == "dash":
                        pass
                    else:
                        vsl_lim   = self._vsl_limit(t_vals)
                        tier1_ind = self._tier1_ind_limit(t_vals)
                        tier1_res = self._tier1_res_limit(t_vals)
                        any_lim   = self._strictest(t_vals)
                        if any_lim is not None:
                            _loq_cmp = loq_map.get(cmp_name)
                            _nd_at_loq = flag_cell == "ND"
                            # GREY first: ND is a non-detection regardless of stored value
                            if _nd_at_loq:
                                if _loq_cmp is not None and _loq_cmp > any_lim:
                                    c.font = _font(val, bold=True)
                                    c.number_format = '"<"0.0##'
                                    has_gray = True
                            # BOLD: threshold < LOD/LOQ → uncertain exclusion
                            elif flag_cell in ("<LOD", "<LOQ", "<"):
                                lod_num = (
                                    lod_cell              if lod_cell              is not None else
                                    lod_map.get(cmp_name) if lod_map.get(cmp_name) is not None else
                                    loq_map.get(cmp_name) if loq_map.get(cmp_name) is not None else
                                    (num_v if isinstance(num_v, (int, float)) else None)
                                )
                                if lod_num is not None and lod_num > any_lim:
                                    c.font = _font(val, bold=True)
                                    has_gray = True
                            # COLOUR: real detection vs threshold
                            elif isinstance(num_v, (int, float)):
                                if tier1_ind is not None and num_v > tier1_ind:
                                    c.fill = PINK
                                    c.font = Font(**FHE, bold=True)
                                elif tier1_res is not None and num_v > tier1_res:
                                    c.fill = L_BLUE
                                    c.font = Font(**FHE, bold=True)
                                elif vsl_lim is not None and num_v > vsl_lim:
                                    c.fill = YELLOW
                                    c.font = Font(**FHE, bold=True)
            data_row += 1

        # ── Merge borehole column cells vertically ───────────────────────
        if has_depth and len(samples) > 1:
            # Walk sample rows and merge runs of the same borehole
            run_bh  = split_map[samples[0]][0]
            run_start = first_sample_row
            for idx, sid in enumerate(samples[1:], 1):
                bh = split_map[sid][0]
                row_num = first_sample_row + idx
                if bh != run_bh:
                    if row_num - 1 > run_start:   # >1 row → merge
                        ws.merge_cells(
                            start_row=run_start, start_column=1,
                            end_row=row_num - 1,   end_column=1
                        )
                        ws.cell(run_start, 1).alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )
                    run_bh    = bh
                    run_start = row_num
            # Flush last run
            last_row = first_sample_row + len(samples) - 1
            if last_row > run_start:
                ws.merge_cells(
                    start_row=run_start, start_column=1,
                    end_row=last_row,     end_column=1
                )
                ws.cell(run_start, 1).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        n_legend = self._write_legend(ws, data_row + 1, include_gray=has_gray,
                                      thresh_keys=thresh_keys)
        # ── Threshold source footnotes (only for keys with ≥1 defined value) ──
        active_keys = [k for k in thresh_keys
                       if any(thresh_vals.get(c, {}).get(k) is not None for c in compounds)]
        note_row = data_row + 1 + n_legend + 1
        for note in self._threshold_source_notes(active_keys):
            c = ws.cell(row=note_row, column=1, value=f"* {note}")
            c.font = Font(**FEN, italic=True, color="808080")
            c.fill = WHITE
            note_row += 1

        # ── Uncertain compound notes ────────────────────────────────────
        for cmp, thresh_name in uncertain_compounds.items():
            if cmp in compounds:
                note_text = (
                    f"* {cmp}: קיים ערך סף לתרכובת דומה בשם: {thresh_name}"
                    if thresh_name else
                    f"* {cmp}: קיים ערך סף לתרכובת דומה — נדרש לבדוק בטבלת ערכי הסף"
                )
                c = ws.cell(row=note_row, column=1, value=note_text)
                c.font = Font(**FEN, italic=True, color="808080")
                c.fill = WHITE
                note_row += 1

        self._auto_width(ws, total_cols)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _thresh_keys(self, atype: str) -> list[str]:
        valid = set(ANALYSIS_THRESHOLDS.get(atype, []))
        if self.sel_thresh is not None:
            # Preserve user's ordering; drop keys not valid for this atype
            return [k for k in self.sel_thresh if k in valid]
        return ANALYSIS_THRESHOLDS.get(atype, [])

    @staticmethod
    def _strictest(t_vals: dict) -> float | None:
        vals = [v for v in t_vals.values() if v is not None]
        return min(vals) if vals else None

    @staticmethod
    def _vsl_limit(t_vals: dict) -> float | None:
        vals = [v for k, v in t_vals.items() if v is not None and _is_vsl_key(k)]
        return min(vals) if vals else None

    @staticmethod
    def _tier1_limit(t_vals: dict) -> float | None:
        """Strictest of all Tier1-type thresholds."""
        vals = [v for k, v in t_vals.items()
                if v is not None and "TIER1" in k]
        return min(vals) if vals else None

    @staticmethod
    def _tier1_ind_limit(t_vals: dict) -> float | None:
        vals = [v for k, v in t_vals.items() if v is not None and _is_ind_key(k)]
        return min(vals) if vals else None

    @staticmethod
    def _tier1_res_limit(t_vals: dict) -> float | None:
        vals = [v for k, v in t_vals.items() if v is not None and _is_res_key(k)]
        return min(vals) if vals else None

    def _write_header_row(self, ws, row_num: int, total_cols: int, hinfo: dict | None = None) -> bool:
        project = (hinfo.get("project", "") if hinfo else self.project) or ""
        client  = (hinfo.get("client",  "") if hinfo else self.client)  or ""
        if not project.strip() and not client.strip():
            return False
        if hinfo:
            parts = [
                ("שם פרויקט:", hinfo.get("project", "")),
                ("תאריך:",     hinfo.get("date", "")),
                ("מזמין:",     hinfo.get("client", "")),
            ]
            span = max(1, total_cols // len(parts))
            for i, (label, val) in enumerate(parts):
                col_start = i * span + 1
                col_end   = (i + 1) * span if i < len(parts) - 1 else total_cols
                ws.merge_cells(start_row=row_num, start_column=col_start,
                               end_row=row_num, end_column=col_end)
                c = ws.cell(row=row_num, column=col_start,
                            value=f"{label}  {val}")
                c.font      = Font(**FHE, bold=True)
                c.alignment = WRAP_C
                c.border    = THIN
        else:
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=total_cols)
            c = ws.cell(row=row_num, column=1,
                        value=f"{self.project}  |  {self.rep_date}  |  {self.client}")
            c.font      = Font(**FHE, bold=True)
            c.alignment = WRAP_C
            c.border    = THIN
        return True

    @staticmethod
    def _threshold_source_notes(thresh_keys: list[str]) -> list[str]:
        """Return unique source footnote strings for the selected threshold keys, in order."""
        seen: set[str] = set()
        out: list[str] = []
        for k in thresh_keys:
            src = _THRESHOLD_SOURCES.get(k)
            if src and src not in seen:
                seen.add(src)
                out.append(src)
        return out

    @staticmethod
    def _write_legend(ws, start_row: int, include_gray: bool = True,
                      thresh_keys: list[str] | None = None) -> int:
        keys = thresh_keys or []
        has_ind = any(_is_ind_key(k) for k in keys)
        has_res = any(_is_res_key(k) for k in keys)
        has_vsl = any(_is_vsl_key(k) for k in keys)
        items = []
        if has_ind:
            items.append(("חריגה מ-Tier 1 תעשייתי", PINK))
        if has_res:
            items.append(("חריגה מ-Tier 1 מגורים",  L_BLUE))
        if has_vsl:
            items.append(("חריגה מערך VSL",          YELLOW))
        if include_gray:
            items.append(("ערכים החורגים מערך סף הגילוי מודגשים", None))
        for i, (label, fill) in enumerate(items):
            c = ws.cell(row=start_row + i, column=1, value=label)
            if fill is not None:
                c.font  = Font(name="David", size=9, bold=True)
                c.fill  = fill
            else:
                # Bold only — no gray fill
                c.font  = Font(name="David", size=9, bold=True)
            c.border = THIN
            c.alignment = Alignment(horizontal="right", vertical="center")
        return len(items)

    @staticmethod
    def _auto_width(ws, n_cols: int, hdr_row: int = 2):
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 13
        for ci in range(3, n_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 13
        ws.row_dimensions[1].height = 20
        for r in range(2, hdr_row):
            ws.row_dimensions[r].height = 20   # metadata rows
        ws.row_dimensions[hdr_row].height = 28  # column header row


# ── Standalone helper: simple KTE groundwater BTEX report (A–H layout) ─────────

def _strip_ns(root: ET.Element) -> ET.Element:
    """Remove XML namespaces for easier searching (for SpreadsheetML .XLS)."""
    xml = ET.tostring(root, encoding="unicode")
    xml = re.sub(r'\s+xmlns(:\w+)?="[^"]*"', "", xml)
    xml = re.sub(r"<(\w+):", "<", xml)
    xml = re.sub(r"</(\w+):", "</", xml)
    xml = re.sub(r"(\s)(\w+):", r"\1", xml)
    return ET.fromstring(xml)


def build_kte_gw_btex_simple_from_xml(
    input_path: str | os.PathLike | bytes | io.BytesIO,
    output_path: str | os.PathLike | io.BytesIO,
) -> str | io.BytesIO:
    """
    Build a simple groundwater BTEX+MTBE sheet (A–H) from
    KTE 'Client GROUNDWATER - 1' SpreadsheetML XML report.

    input_path  — file path, raw bytes, or BytesIO
    output_path — file path or BytesIO
    """
    if isinstance(input_path, (bytes, bytearray)):
        raw = input_path
    elif isinstance(input_path, io.BytesIO):
        raw = input_path.read()
    else:
        raw = Path(input_path).read_bytes()
    root = ET.fromstring(raw)
    root = _strip_ns(root)

    # Find "Client GROUNDWATER - 1" worksheet
    ws_xml = None
    for w in root.findall(".//Worksheet"):
        name = (w.get("Name") or "").strip().lower()
        if "client groundwater - 1" in name:
            ws_xml = w
            break
    if ws_xml is None:
        raise RuntimeError("לא נמצא worksheet בשם 'Client GROUNDWATER - 1'")

    table = ws_xml.find(".//Table")
    if table is None:
        raise RuntimeError("לא נמצא <Table> ב‑Client GROUNDWATER - 1")

    # Convert XML rows → list[list[str]]
    rows: list[list[str]] = []
    for row_el in table.findall("Row"):
        cells: list[str] = []
        prev_idx = 0
        for cell_el in row_el.findall("Cell"):
            idx_attr = cell_el.get("Index")
            if idx_attr is not None:
                idx = int(idx_attr)
                gap = idx - 1 - prev_idx
                if gap > 0:
                    cells.extend([""] * gap)
            data_el = cell_el.find("Data")
            val = data_el.text if (data_el is not None and data_el.text) else ""
            cells.append(val.strip())
            prev_idx = len(cells) - 1
        rows.append(cells)

    # Locate sample IDs and sampling dates
    sample_row = date_row = None
    for r in rows:
        joined = " ".join(r).lower()
        if "client sample id" in joined:
            sample_row = r
        if "client sampling date" in joined:
            date_row = r
    if sample_row is None or date_row is None:
        raise RuntimeError("לא מצאתי שורות Client Sample ID / Client Sampling Date")

    # In this SpreadsheetML export the data starts at column index 4 (E in Excel)
    well_names = sample_row[4:]
    dates = date_row[4:]

    def _find_param_row(name_substr: str, preferred_unit: str | None = None) -> list[str] | None:
        """
        Find first row whose first cell contains name_substr (case-insensitive).
        If preferred_unit is given, prefer a row whose Unit column matches it.
        """
        key = name_substr.lower()
        candidates: list[list[str]] = []
        for r in rows:
            if r and key in r[0].lower():
                candidates.append(r)
        if not candidates:
            return None
        if preferred_unit is None:
            return candidates[0]
        for r in candidates:
            if len(r) > 2 and (r[2] or "").strip() == preferred_unit:
                return r
        return candidates[0]

    r_benzene = _find_param_row("benzene")
    r_toluene = _find_param_row("toluene")
    r_ethylbenz = _find_param_row("ethylbenzene")
    r_meta_para = _find_param_row("meta- & para-xylene")
    r_ortho_xyl = _find_param_row("ortho-xylene")
    r_mtbe = _find_param_row("methyl tert-butyl ether (mtbe)")

    if not all([r_benzene, r_toluene, r_ethylbenz, r_meta_para, r_ortho_xyl, r_mtbe]):
        raise RuntimeError("לא נמצאו כל שורות BTEX/MTBE ב‑INPUT")

    def _vals_ugL_to_mgL(r: list[str]) -> list[float | str]:
        unit = (r[2] or "").strip()  # col C = Unit
        factor = 0.001 if unit == "µg/L" else 1.0
        out: list[float | str] = []
        for v in r[4:]:
            v = v.strip()
            if not v:
                out.append("")
            elif v.startswith("<"):
                try:
                    num = float(v[1:])
                    out.append(f"<{round(num * factor, 4)}")
                except ValueError:
                    out.append(v)
            else:
                try:
                    num = float(v)
                    out.append(round(num * factor, 4))
                except ValueError:
                    out.append(v)
        return out

    benzene_vals = _vals_ugL_to_mgL(r_benzene)
    toluene_vals = _vals_ugL_to_mgL(r_toluene)
    ethyl_vals = _vals_ugL_to_mgL(r_ethylbenz)
    meta_para_vals = _vals_ugL_to_mgL(r_meta_para)
    ortho_vals = _vals_ugL_to_mgL(r_ortho_xyl)
    mtbe_vals = _vals_ugL_to_mgL(r_mtbe)

    # Xylene = sum(meta+para, ortho)
    xylene_vals: list[float | str] = []
    for mp, o in zip(meta_para_vals, ortho_vals):
        try:
            mp_num = float(str(mp).lstrip("<"))
            o_num = float(str(o).lstrip("<"))
            xylene_vals.append(round(mp_num + o_num, 4))
        except ValueError:
            xylene_vals.append("")

    # Build simple Excel sheet
    wb = openpyxl.Workbook()

    # ── Sheet 1: BTEX + MTBE ─────────────────────────────────────────────
    ws_out = wb.active
    ws_out.title = "מי תהום BTEX"
    ws_out.sheet_view.rightToLeft = True

    NUM_FMT = "0.0000"   # 4 decimal places for all numeric data cells

    def _set_cell(r: int, c: int, val, bold: bool = False, num: bool = False):
        cell = ws_out.cell(row=r, column=c, value=val)
        cell.font = Font(**FHE, bold=bold)   # David 9 everywhere
        cell.alignment = CENTER
        cell.border = THIN
        if num and val not in (None, ""):
            cell.number_format = NUM_FMT
        return cell

    # Row 1: headers (all David 9 bold)
    _set_cell(1, 1, "שם קידוח", bold=True)
    _set_cell(1, 2, "תאריך דיגום", bold=True)
    _set_cell(1, 3, "", bold=True)
    _set_cell(1, 4, "בנזן", bold=True)
    _set_cell(1, 5, "טולואן", bold=True)
    _set_cell(1, 6, "אתיל בנזן", bold=True)
    _set_cell(1, 7, "כסילן", bold=True)
    # MTBE is English → Times New Roman 9 bold
    _c = ws_out.cell(row=1, column=8, value="MTBE")
    _c.font = Font(**FEN, bold=True)
    _c.alignment = CENTER
    _c.border = THIN

    # Row 2: units  (A,B empty | C יחידות | D-H מ"ג/ליטר)
    _set_cell(2, 1, "")
    _set_cell(2, 2, "")
    _set_cell(2, 3, "יחידות", bold=True)
    for col in range(4, 9):
        _set_cell(2, col, 'מ"ג/ליטר')

    # Row 3: restoration targets  (A,B empty | C label | D-H values David 9 + 4dp)
    _set_cell(3, 1, "")
    _set_cell(3, 2, "")
    _set_cell(3, 3, "ערכי יעד לשיקום ^", bold=True)
    targets = {
        4: 0.094,   # Benzene
        5: 13.0,    # Toluene
        6: 5.6,     # Ethylbenzene
        7: 9.4,     # Xylenes
        8: 0.75,    # MTBE
    }
    for col in range(4, 9):
        _set_cell(3, col, targets.get(col, ""), num=True)

    def _norm_well(name: str) -> str:
        """MT-x → מת-x"""
        s = (name or "").strip()
        if s.upper().startswith("MT-"):
            return "מת-" + s[3:]
        return s

    # Sample rows (rows 4+)
    for idx, (well, dt) in enumerate(zip(well_names, dates), start=4):
        # well name: Hebrew → David 9, English → Times 8 (auto-detect via _font)
        _wn = _norm_well(well)
        _c = ws_out.cell(row=idx, column=1, value=_wn)
        _c.font = _font(_wn)
        _c.alignment = CENTER
        _c.border = THIN
        _set_cell(idx, 2, dt or "")           # date      — David 9
        _set_cell(idx, 3, "")

        def _put(col: int, arr: list[float | str]):
            pos = idx - 4
            if 0 <= pos < len(arr):
                val = arr[pos]
                cell = _set_cell(idx, col, val, num=True)   # David 9 + 4dp
                t = targets.get(col)
                if t is not None:
                    try:
                        num_val = float(str(val).lstrip("<"))
                        if num_val > t:
                            cell.fill = GRAY
                            cell.font = Font(**FHE, bold=True)
                    except (ValueError, TypeError):
                        pass

        _put(4, benzene_vals)
        _put(5, toluene_vals)
        _put(6, ethyl_vals)
        _put(7, xylene_vals)
        _put(8, mtbe_vals)

    # Basic widths
    ws_out.column_dimensions["A"].width = 18
    ws_out.column_dimensions["B"].width = 14
    ws_out.column_dimensions["C"].width = 18
    for col in range(4, 9):
        ws_out.column_dimensions[get_column_letter(col)].width = 12

    # ── Sheet 2: Chlorite/Chlorate/Chloride/Perchlorate ─────────────────
    # Parameters:
    #   D: כלוריט + כלוראט  → "Sum of chlorites and chlorates"
    #   E: כלוריד           → "Chloride" (mg/L row)
    #   F: פרכלורט          → "Perchlorate"
    #   G: כלוראט           → "Chlorate"

    r_sum_chlor = _find_param_row("sum of chlorites and chlorates")
    r_chloride  = _find_param_row("chloride", preferred_unit="mg/L")
    r_percl     = _find_param_row("perchlorate")
    r_chlorate  = _find_param_row("chlorate")

    def _save_wb(wb, dest):
        if isinstance(dest, io.BytesIO):
            wb.save(dest)
            dest.seek(0)
            return dest
        os.makedirs(os.path.dirname(str(dest)) or ".", exist_ok=True)
        wb.save(dest)
        return str(dest)

    if any(r is None for r in (r_sum_chlor, r_chloride, r_percl, r_chlorate)):
        return _save_wb(wb, output_path)

    def _vals_to_mgL(r: list[str]) -> list[float | str]:
        unit = (r[2] or "").strip()
        factor = 0.001 if unit == "µg/L" else 1.0
        out: list[float | str] = []
        for v in r[4:]:
            v = v.strip()
            if not v:
                out.append("")
            elif v.startswith("<"):
                try:
                    num = float(v[1:])
                    out.append(f"<{round(num * factor, 4)}")
                except ValueError:
                    out.append(v)
            else:
                try:
                    num = float(v)
                    out.append(round(num * factor, 4))
                except ValueError:
                    out.append(v)
        return out

    sum_chlor_vals = _vals_to_mgL(r_sum_chlor)
    chloride_vals  = _vals_to_mgL(r_chloride)
    percl_vals     = _vals_to_mgL(r_percl)
    chlorate_vals  = _vals_to_mgL(r_chlorate)

    ws2 = wb.create_sheet(title="מי שתייה – כלורידים")
    ws2.sheet_view.rightToLeft = True

    def _set2(r: int, c: int, val, bold: bool = False, he: bool = True):
        cell = ws2.cell(row=r, column=c, value=val)
        if he:
            cell.font = Font(**FHE, bold=bold)
        else:
            base = FEN.copy()
            base["bold"] = bold
            cell.font = Font(**base)
        cell.alignment = CENTER
        cell.border = THIN
        return cell

    # Row 1: headers
    _set2(1, 1, "שם קידוח", bold=True)
    _set2(1, 2, "תאריך דיגום", bold=True)
    _set2(1, 3, "", bold=True)
    _set2(1, 4, "כלוריט + כלוראט", bold=True)
    _set2(1, 5, "כלוריד", bold=True)
    _set2(1, 6, "פרכלורט", bold=True)
    _set2(1, 7, "כלוראט", bold=True)

    # Row 2: units
    _set2(2, 1, "")
    _set2(2, 2, "")
    _set2(2, 3, "יחידות", bold=True)
    for col in range(4, 8):
        _set2(2, col, 'מ"ג/ליטר')

    # Row 3: drinking water thresholds
    _set2(3, 1, "")
    _set2(3, 2, "")
    _set2(3, 3, "ערך סף למי שתייה^", bold=True)
    _set2(3, 4, "--")
    _set2(3, 5, 400)   # כלוריד
    _set2(3, 6, "--")
    _set2(3, 7, "--")

    # Sample rows
    for idx, (well, dt) in enumerate(zip(well_names, dates), start=4):
        _set2(idx, 1, well or "")
        _set2(idx, 2, dt or "")
        _set2(idx, 3, "")

        def _put2(col: int, arr: list[float | str]):
            pos = idx - 4
            if 0 <= pos < len(arr):
                _set2(idx, col, arr[pos])

        _put2(4, sum_chlor_vals)
        _put2(5, chloride_vals)
        _put2(6, percl_vals)
        _put2(7, chlorate_vals)

    # Column widths
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 22
    for col in range(4, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    return _save_wb(wb, output_path)
